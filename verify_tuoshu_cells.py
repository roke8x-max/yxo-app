import urllib.request, json, io, zipfile
from openpyxl import load_workbook

BASE = "http://127.0.0.1:5011"

def call(path, data):
    req = urllib.request.Request(
        BASE + path,
        data=json.dumps(data).encode("utf-8"),
        headers={"Content-Type": "application/json"},
    )
    return urllib.request.urlopen(req, timeout=60).read()

base = {
    "user": "毛骁洋",
    "date_from": "2026-08-01",
    "date_to": "2026-08-05",
    "group_by": "train_dest",
    "booking_date": "2026-08-03",
}
meta = json.loads(call("/api/tuoshu/preview", base))
groups = meta.get("groups", [])
print("PREVIEW total=%s boxes=%s unresolved=%s ngroups=%s" % (
    meta.get("total"), meta.get("boxes"), meta.get("unresolved"), len(groups)))

picked = [{"train_no": g["train_no"], "departure_date": g["departure_date"], "station_cn": g["station_cn"]} for g in groups]
raw = call("/api/tuoshu/generate", {"user": "毛骁洋", "picked": picked})
print("GEN bytes=%d" % len(raw))

z = zipfile.ZipFile(io.BytesIO(raw))
names = z.namelist()
print("ZIP files=%d" % len(names))
for name in names[:6]:
    wb = load_workbook(io.BytesIO(z.read(name)))
    ws = wb.active
    print("== %s" % name)
    print("   E4 订舱日期 :", repr(ws["E4"].value))
    print("   E13 发运    :", repr(ws["E13"].value))
    print("   E15 目的站  :", repr(ws["E15"].value))
    print("   B24 箱量    :", repr(ws["B24"].value))

# 单份直出校验：取一个可解析目的站的组
single = None
for g in groups:
    if "未知" not in g.get("station_cn", ""):
        single = g
        break
if single:
    sp = [{"train_no": single["train_no"], "departure_date": single["departure_date"], "station_cn": single["station_cn"]}]
    raw2 = call("/api/tuoshu/generate", {"user": "毛骁洋", "picked": sp})
    wb = load_workbook(io.BytesIO(raw2))
    ws = wb.active
    print("== SINGLE %s / %s / %s" % (single["train_no"], single["departure_date"], single["station_cn"]))
    print("   E4 订舱日期 :", repr(ws["E4"].value))
    print("   E13 发运    :", repr(ws["E13"].value))
    print("   E15 目的站  :", repr(ws["E15"].value))
    print("   B24 箱量    :", repr(ws["B24"].value))
print("DONE")
