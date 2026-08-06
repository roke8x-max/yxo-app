#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
YXO 订舱数据管理（方案 A 正式版）
---------------------------------------------------
- 后端: Flask + SQLite（单数据源，脚本可读写）
- 前端: 浏览器表格（Excel/飞书式多条件筛选、下拉选项、看板标签、统计、序号、个人筛选隔离）
- 访问: 本机 http://localhost:5011  （经 nginx 反代 http://公网IP:5000/yxo/）
- 数据: 首次启动从 config.IMPORT_FILE 导入
启动: 双击 start.bat
"""
import os
import re
import json
import sqlite3
import socket
from datetime import datetime

from flask import Flask, request, jsonify, render_template, send_file

import config
from import_excel import run_import
from admin_api import admin_bp

BASE = config.BASE
DB = config.DB_PATH

app = Flask(__name__)
app.register_blueprint(admin_bp)   # 管理页（仅毛骁洋）：/admin + /api/admin/*

# 莫斯科 / 明斯克 回退站点
_MOSCOW_FB = {"别雷拉斯特", "沃尔西诺", "电煤", "谢利亚季诺"}
_MINSK_FB = {"科里亚季奇"}


# ====================== 数据库 ======================
def get_db():
    conn = sqlite3.connect(DB)
    conn.row_factory = sqlite3.Row
    return conn


def default_values():
    """新增行时空单元格套用的默认值"""
    return {f["name"]: f["default"] for f in config.FIELD_DEFS if f.get("default")}


def init_options(conn):
    """选项表：维护各单选字段的选项（可在界面增删）。以 config 为种子，仅补充缺失项。"""
    conn.execute("""CREATE TABLE IF NOT EXISTS field_options (
        field TEXT PRIMARY KEY, options TEXT)""")
    for f in config.FIELD_DEFS:
        if f.get("options") is not None:
            field = f["name"]
            if not conn.execute("SELECT 1 FROM field_options WHERE field=?", (field,)).fetchone():
                conn.execute("INSERT INTO field_options(field, options) VALUES(?, ?)",
                             (field, json.dumps(f["options"], ensure_ascii=False)))


def migrate_db(conn, cur):
    """历史字段改名 / 选项清理（在补全缺失列之前执行）"""
    if "数据状态" in cur and "状态" not in cur:
        conn.execute('ALTER TABLE records RENAME COLUMN "数据状态" TO "状态"')
    try:
        conn.execute('UPDATE records SET "状态"=\'正常\' WHERE "状态"=\'待定\'')
    except Exception:
        pass


def fill_defaults_once(conn):
    """升级时一次性把"已存在但为空"的单元格补上默认值（仅一次，不覆盖用户后续清空）"""
    done = conn.execute("SELECT value FROM meta_kv WHERE key='defaults_v1'").fetchone()
    if done:
        return
    for f in config.FIELD_DEFS:
        d = f.get("default")
        if d:
            conn.execute(
                f'UPDATE records SET "{f["name"]}"=? WHERE "{f["name"]}" IS NULL OR TRIM("{f["name"]}")=?',
                (d, ""))
    conn.execute(
        "INSERT INTO meta_kv(key, value) VALUES('defaults_v1','1') "
        "ON CONFLICT(key) DO UPDATE SET value='1'")


def backfill_ledger_month(conn):
    """升级时一次性回填“台账月份”：默认 = 发班时间所在月（YYYY-MM），便于在月份标签里归属。
    仅回填空值，不覆盖用户后续手动调整。"""
    done = conn.execute("SELECT value FROM meta_kv WHERE key='ledger_month_v1'").fetchone()
    if done:
        return
    rows = conn.execute('SELECT id, "发班时间" FROM records WHERE "台账月份" IS NULL OR TRIM("台账月份")=""').fetchall()
    for r in rows:
        s = r["发班时间"] or ""
        mk = s[:7].replace("/", "-") if s else ""
        if re.match(r"^\d{4}-\d{2}$", mk):
            conn.execute('UPDATE records SET "台账月份"=? WHERE id=?', (mk, r["id"]))
    conn.execute(
        "INSERT INTO meta_kv(key, value) VALUES('ledger_month_v1','1') "
        "ON CONFLICT(key) DO UPDATE SET value='1'")


def init_db():
    os.makedirs(config.DATA_DIR, exist_ok=True)
    conn = get_db()
    cols = ", ".join([f'"{f}" TEXT' for f in config.ALL_FIELDS])
    conn.execute(f"""
        CREATE TABLE IF NOT EXISTS records (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            seq INTEGER,
            {cols},
            updated_at TEXT, updated_by TEXT
        )
    """)
    cur = {r[1] for r in conn.execute("PRAGMA table_info(records)").fetchall()}
    # 先处理改名（必须在补全缺失列之前，否则"状态"会被误建）
    migrate_db(conn, cur)
    cur = {r[1] for r in conn.execute("PRAGMA table_info(records)").fetchall()}
    # 配置变更时补全缺失列（备注等新增字段）
    for f in config.ALL_FIELDS:
        if f not in cur:
            conn.execute(f'ALTER TABLE records ADD COLUMN "{f}" TEXT')
    # 手动排序字段（支持插入到某行上方/下方）
    if "order_idx" not in cur:
        conn.execute('ALTER TABLE records ADD COLUMN order_idx REAL')
    conn.execute("UPDATE records SET order_idx = seq WHERE order_idx IS NULL")
    # 软删除（回收站）：删除只打标记，可在管理页恢复
    if "is_deleted" not in cur:
        conn.execute('ALTER TABLE records ADD COLUMN is_deleted INTEGER DEFAULT 0')
    if "deleted_at" not in cur:
        conn.execute('ALTER TABLE records ADD COLUMN deleted_at TEXT')
    if "deleted_by" not in cur:
        conn.execute('ALTER TABLE records ADD COLUMN deleted_by TEXT')
    # 常用查询列建索引：客服机器人按 班列号/客户编码/箱号 查询、网页筛选、按发班时间统计都受益。
    # 班列号跨年可能重复，故用普通索引（不吃 UNIQUE）。
    for _idx in (
        'CREATE INDEX IF NOT EXISTS idx_records_train ON records("班列号")',
        'CREATE INDEX IF NOT EXISTS idx_records_code ON records("客户编码")',
        'CREATE INDEX IF NOT EXISTS idx_records_box ON records("箱号")',
        'CREATE INDEX IF NOT EXISTS idx_records_depart ON records("发班时间")',
        'CREATE INDEX IF NOT EXISTS idx_records_company ON records("开票子公司名称")',
    ):
        try:
            conn.execute(_idx)
        except Exception:
            pass
    # 专列手动状态（未发运/已发运/已完成）：以 (班列号, 发班年份) 为键，人工维护，不与箱级 records 混表。
    conn.execute("""CREATE TABLE IF NOT EXISTS train_meta (
        train TEXT NOT NULL,
        year TEXT NOT NULL,
        train_status TEXT NOT NULL,
        updated_at TEXT,
        PRIMARY KEY (train, year)
    )""")
    # P3 运踪日志本地表（B方案）：替代飞书 TABLE_LOG，成为运踪转发日志唯一数据源。
    # 粒度=「一次转发事件」；运踪邮件只有班列号、无箱号，故不挂 records，独立建表。
    # 关联/去重键 = 邮件唯一标识 + 接收公司（同一封运踪邮件会按路由转发给多家，每家一条）。
    conn.execute("""CREATE TABLE IF NOT EXISTS tracing_log (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        log_id TEXT,
        train_no TEXT,
        train_key TEXT,
        company TEXT,
        mail_msg_id TEXT NOT NULL,
        forward_detail TEXT,
        log_date TEXT,
        created_at TEXT DEFAULT (datetime('now','localtime')),
        UNIQUE(mail_msg_id, company)
    )""")
    conn.execute("CREATE INDEX IF NOT EXISTS idx_tracing_train ON tracing_log(train_no)")
    conn.execute("CREATE INDEX IF NOT EXISTS idx_tracing_date ON tracing_log(log_date)")
    # 体检D（2026-08-04）：运踪快照表 —— 补系统的"时间维度"。
    # 粒度=「一次运踪事件」；纯增量写入，不影响 tracing_log 现有逻辑。
    # 幂等键 (train_key, event_time, source)，避免同一封邮件按扇形转发多家时重复记。
    conn.execute("""CREATE TABLE IF NOT EXISTS tracing_snapshot (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        train_key TEXT NOT NULL,
        box_no TEXT,
        node TEXT,
        status TEXT,
        event_time TEXT NOT NULL,
        source TEXT,
        created_at TEXT DEFAULT (datetime('now','localtime')),
        UNIQUE(train_key, event_time, source)
    )""")
    conn.execute("CREATE INDEX IF NOT EXISTS idx_snap_train ON tracing_snapshot(train_key, event_time)")
    # 舱单统一导入（#199）：写库留痕 + 批次 + 整批回退 + 快照还原。
    # update_log：每一次「写」的明细（含旧值），支撑单条撤销 / 复制 TSV / 审计。
    conn.execute("""CREATE TABLE IF NOT EXISTS update_log (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        batch_id TEXT NOT NULL,
        batch_type TEXT,
        record_id INTEGER,
        "客户编码" TEXT,
        "箱号" TEXT,
        field TEXT,
        old_value TEXT,
        new_value TEXT,
        action TEXT,
        source_file TEXT,
        operator TEXT,
        reverted INTEGER DEFAULT 0,
        reverted_at TEXT,
        created_at TEXT
    )""")
    conn.execute("CREATE INDEX IF NOT EXISTS idx_ul_batch ON update_log(batch_id)")
    conn.execute("CREATE INDEX IF NOT EXISTS idx_ul_rec ON update_log(record_id)")
    # import_batch：一次导入操作的批次元信息 + 整库快照路径（核弹级还原用）。
    conn.execute("""CREATE TABLE IF NOT EXISTS import_batch (
        batch_id TEXT PRIMARY KEY,
        batch_type TEXT,
        source_files TEXT,
        snapshot TEXT,
        n_update INTEGER DEFAULT 0,
        n_insert INTEGER DEFAULT 0,
        n_alert INTEGER DEFAULT 0,
        operator TEXT,
        reverted INTEGER DEFAULT 0,
        created_at TEXT
    )""")
    init_options(conn)
    fill_defaults_once(conn)
    backfill_ledger_month(conn)
    conn.commit()
    # 首次为空则导入
    cnt = conn.execute("SELECT COUNT(*) AS c FROM records").fetchone()["c"]
    if cnt == 0:
        run_import(conn)
    conn.close()


# ====================== 价格计算（复用 price_config.json） ======================
def load_price_config():
    try:
        # utf-8-sig：兼容记事本另存为带来的 BOM（体检报告 3.6）
        with open(config.PRICE_CONFIG, "r", encoding="utf-8-sig") as f:
            return json.load(f)
    except Exception:
        return {"prices": {}, "exrates": {}}


def compute_price(port, dest, cntr_type, depart_str, cfg):
    if not (port and dest):
        return None
    ct = "COC" if cntr_type and "COC" in str(cntr_type).upper() else "SOC"
    mk = None
    if depart_str:
        # 只取日期部分（兼容 "2026/08/08"、"2026-08-08"、"2026/8/8 00:00:00" 等）
        head = str(depart_str).strip().split(" ")[0].split("T")[0]
        for fmt in ("%Y/%m/%d", "%Y-%m-%d"):
            try:
                mk = datetime.strptime(head, fmt).strftime("%Y-%m")
                break
            except Exception:
                continue
    if not mk:
        return None
    prices = cfg.get("prices", {}).get(mk, {})
    if not prices:
        return None
    key = f"{port}|{dest}"
    pi = prices.get(key)
    if pi is None and dest in _MOSCOW_FB:
        pi = prices.get(f"{port}|莫斯科")
    if pi is None and dest in _MINSK_FB:
        pi = prices.get(f"{port}|明斯克")
    if pi is None:
        return None
    rp = pi.get(ct)
    if rp is None and ct == "COC":
        rp = pi.get("SOC")
    if rp is None:
        return None
    cur = pi.get("currency", "USD")
    if cur == "RMB":
        return round(rp, 2)
    rate = cfg.get("exrates", {}).get(mk, {}).get("USD")
    if rate is None:
        return None
    return round(rp * rate, 2)


# ====================== 数据版本号（多人协同：他人改动自动刷新） ======================
def bump_version(conn):
    conn.execute(
        "INSERT INTO meta_kv (key, value) VALUES ('data_version', '1') "
        "ON CONFLICT(key) DO UPDATE SET value = CAST(CAST(value AS INTEGER) + 1 AS TEXT)")


def get_version(conn):
    row = conn.execute("SELECT value FROM meta_kv WHERE key='data_version'").fetchone()
    return int(row["value"]) if row else 0


# ====================== API ======================
@app.route("/")
def index():
    from flask import make_response
    resp = make_response(render_template("index.html"))
    resp.headers["Cache-Control"] = "no-store, no-cache, must-revalidate, max-age=0"
    resp.headers["Pragma"] = "no-cache"
    resp.headers["Expires"] = "0"
    return resp


@app.route("/api/meta")
def api_meta():
    conn = get_db()
    opt_map = {}
    for row in conn.execute("SELECT field, options FROM field_options"):
        try:
            opt_map[row["field"]] = json.loads(row["options"])
        except Exception:
            pass
    conn.close()
    return jsonify({
        "columns": [
            {
                "name": f["name"], "kind": f["kind"], "editable": True,
                "label": f.get("label"),
                "type": f.get("type", "text"),
                "options": opt_map.get(f["name"], f.get("options")),
                "hint": f.get("hint"),
                "default": f.get("default"),
                "free_text": bool(f.get("free_text")),
                "maintainable": bool(f.get("maintainable")),
                "trainTypes": f.get("trainTypes"),
            } for f in config.FIELD_DEFS
        ],
        "users": config.USERS,
        "company_field": config.COMPANY_FIELD,
        "groupable": config.GROUPABLE,
        "followup_fields": config.FOLLOWUP_FIELDS,
        "manifest_admins": [config.MANIFEST_ADMIN],
    })


@app.route("/api/field_options", methods=["GET", "POST"])
def api_field_options():
    maintainable = {f["name"] for f in config.FIELD_DEFS if f.get("maintainable")}
    if request.method == "GET":
        conn = get_db()
        label_map = {f["name"]: f.get("label") for f in config.FIELD_DEFS}
        out = []
        for row in conn.execute("SELECT field, options FROM field_options"):
            f = row["field"]
            if f in maintainable:
                try:
                    out.append({"field": f, "label": label_map.get(f) or f,
                                "options": json.loads(row["options"])})
                except Exception:
                    pass
        conn.close()
        return jsonify(out)
    # POST：更新某字段的选项（全量写入）
    data = request.get_json(force=True, silent=True) or {}
    field = data.get("field")
    options = data.get("options")
    if field not in maintainable:
        return jsonify({"ok": False, "msg": "该字段不可维护"}), 400
    if not isinstance(options, list):
        return jsonify({"ok": False, "msg": "选项必须是列表"}), 400
    clean = [str(o).strip() for o in options if str(o).strip()]
    conn = get_db()
    conn.execute(
        "INSERT INTO field_options(field, options) VALUES(?, ?) "
        "ON CONFLICT(field) DO UPDATE SET options=excluded.options",
        (field, json.dumps(clean, ensure_ascii=False)))
    bump_version(conn)
    conn.commit()
    conn.close()
    return jsonify({"ok": True})


@app.route("/api/version")
def api_version():
    conn = get_db()
    v = get_version(conn)
    conn.close()
    return jsonify({"version": v})


@app.route("/api/train_summary")
def api_train_summary():
    """专列按「班列号 + 发班年份」聚合（Dashboard 与微信机器人共用同一份数据）。
    退舱箱不计入统计；异常箱 = 箱号为空 或 报放单≠已上传（非退舱）的箱子。"""
    from collections import defaultdict
    conn = get_db()
    cols = '"班列号","发班时间","开票子公司名称","箱属","箱号","报放单","状态","专列状态","备注","口岸"'
    rows = conn.execute(
        f'SELECT {cols} FROM records '
        f'WHERE COALESCE(is_deleted,0)=0 AND "班列类型"=\'专列\'').fetchall()
    meta_rows = conn.execute('SELECT train, year, train_status FROM train_meta').fetchall()
    conn.close()
    meta_map = {(r["train"], r["year"]): r["train_status"] for r in meta_rows}
    trains = defaultdict(list)
    for r in rows:
        train = (r["班列号"] or "").strip()
        dep = r["发班时间"] or ""
        year = dep[:4] if len(dep) >= 4 and dep[:4].isdigit() else "未知"
        trains[(train, year)].append(dict(r))
    today = datetime.now().strftime("%Y-%m-%d")
    out = []
    for (train, year), recs in trains.items():
        # 退舱箱不参与任何统计（设计稿 activeBoxes 口径）
        active = [r for r in recs if (r["状态"] or "") != "退舱"]
        total = len(active)
        tuicang = len(recs) - total
        soc = sum(1 for r in active if (r["箱属"] or "") == "SOC")
        coc = sum(1 for r in active if (r["箱属"] or "") == "COC")
        box_done = sum(1 for r in active if (r["箱号"] or "").strip())
        fang_done = sum(1 for r in active if (r["报放单"] or "") == "已上传")
        dep0 = next((r["发班时间"] for r in recs if (r["发班时间"] or "").strip()), "")
        company = next((r["开票子公司名称"] for r in recs if (r["开票子公司名称"] or "").strip()), "")
        port = next((r["口岸"] for r in recs if (r["口岸"] or "").strip()), "")
        # 已发运 = 发班时间已填且不晚于今天；填了未来日期的算「待发班」
        shipped = bool(dep0.strip()) and dep0[:10] <= today
        # 手动班列状态（未发运/已发运/已完成）：有手动值优先，否则回退自动发运判断
        manual = meta_map.get((train, year))
        status = manual if manual else ("已发运" if shipped else "未发运")
        anomalies = []
        # 专列异常仅认「专列状态=非正常」；箱号空/报放未上传属独立统计卡，不进异常面板。
        # 遍历全部记录（含退舱）：退舱箱仍计入统计排除，但专列状态=非正常的也进异常面板。
        for r in recs:
            status_bad = (r["专列状态"] or "") == "非正常"
            if status_bad:
                anomalies.append({
                    "箱号": r["箱号"] or "",
                    "箱号状态": "未确认" if not (r["箱号"] or "").strip() else "已填",
                    "报放状态": r["报放单"] or "未出",
                    "记录状态": r["专列状态"] or "",
                    "备注": r["备注"] or "",
                })
        out.append({
            "train": train, "year": year, "发班时间": dep0,
            "status": status,
            "负责公司": company, "口岸": port,
            "total": total, "soc": soc, "coc": coc, "退舱": tuicang,
            "box_done": box_done, "box_total": total,
            "fang_done": fang_done, "fang_total": total,
            "报放待跟进": total - fang_done,
            "发运": shipped,
            "异常数": len(anomalies), "有异常": len(anomalies) > 0,
            "anomalies": anomalies,
        })
    out.sort(key=lambda t: (t["发班时间"] or ""), reverse=True)
    return jsonify(out)


@app.route("/api/train_status", methods=["POST"])
def api_train_status():
    """手动维护专列状态（未发运/已发运/已完成），按 (班列号, 发班年份) 存储。"""
    data = request.get_json(silent=True) or {}
    train = (data.get("train") or "").strip()
    year = (data.get("year") or "").strip()
    status = data.get("status")
    if not train:
        return jsonify(ok=False, msg="班列号缺失"), 400
    if status not in ("未发运", "已发运", "已完成"):
        return jsonify(ok=False, msg="状态非法（仅限 未发运/已发运/已完成）"), 400
    conn = get_db()
    conn.execute(
        'INSERT INTO train_meta (train, year, train_status, updated_at) VALUES (?,?,?,?) '
        'ON CONFLICT(train, year) DO UPDATE SET train_status=excluded.train_status, updated_at=excluded.updated_at',
        (train, year, status, datetime.now().isoformat(timespec="seconds")))
    conn.commit()
    conn.close()
    return jsonify(ok=True)


# 导航页用的公开服务状态（无需登录）：仅返回 名称+存活，不暴露端口/路径。
# 个人媒体服务（AList/Navidrome/Jellyfin）的状态在此展示，但其可点链接不挂公网（数科部要求 VPN 内网访问）。
_NAV_SERVICES = [
    ("YXO 订舱数据管理", 5011),
    ("订舱助手 (企微机器人)", 5001),
    ("FileBrowser 文件", 8080),
    ("Netdata 监控", 19999),
    ("AList 网盘", 5244),
    ("Navidrome 音乐", 4533),
    ("Jellyfin 影视", 8096),
]


@app.route("/api/service_status")
def api_service_status():
    out = []
    for name, port in _NAV_SERVICES:
        alive = False
        try:
            with socket.create_connection(("127.0.0.1", port), timeout=1.5):
                alive = True
        except Exception:
            alive = False
        out.append({"name": name, "alive": alive})
    return jsonify({"ok": True, "services": out})


@app.route("/api/rows")
def api_rows():
    conn = get_db()
    flds = ", ".join([f'"{f}"' for f in config.ALL_FIELDS])
    rows = conn.execute(
        f'SELECT id, seq, order_idx, {flds}, updated_at, updated_by FROM records '
        f'WHERE COALESCE(is_deleted,0)=0 ORDER BY order_idx, id'
    ).fetchall()
    conn.close()
    return jsonify([dict(r) for r in rows])


@app.route("/api/row", methods=["POST"])
def api_add_row():
    conn = get_db()
    seq = conn.execute("SELECT COALESCE(MAX(seq),0)+1 AS s FROM records").fetchone()["s"]
    cols = config.ALL_FIELDS
    ph = ", ".join(["?"] * len(cols))
    col_sql = ", ".join([f'"{c}"' for c in cols])
    defaults = default_values()
    vals = [defaults.get(c, "") for c in cols]
    conn.execute(f"INSERT INTO records (seq, order_idx, {col_sql}) VALUES (?, ?, {ph})",
                 [seq, float(seq)] + vals)
    new_id = conn.execute("SELECT last_insert_rowid() AS i").fetchone()["i"]
    bump_version(conn)
    conn.commit()
    conn.close()
    return jsonify({"ok": True, "id": new_id, "seq": seq})


@app.route("/api/row/insert", methods=["POST"])
def api_insert_row():
    data = request.get_json(force=True, silent=True) or {}
    ref_id = data.get("ref_id")
    position = data.get("position", "end")  # before | after | end
    conn = get_db()
    rows = conn.execute("SELECT id, order_idx FROM records ORDER BY order_idx, id").fetchall()
    max_seq = conn.execute("SELECT COALESCE(MAX(seq),0)+1 AS s FROM records").fetchone()["s"]
    if not rows:
        new_order = 1.0
    elif position == "end" or ref_id is None:
        new_order = (rows[-1]["order_idx"] or 0) + 1.0
    else:
        ids = [r["id"] for r in rows]
        try:
            idx = ids.index(ref_id)
        except ValueError:
            new_order = (rows[-1]["order_idx"] or 0) + 1.0
        else:
            if position == "before":
                new_order = (rows[idx]["order_idx"] or 0) - 1.0 if idx == 0 \
                    else ((rows[idx - 1]["order_idx"] or 0) + (rows[idx]["order_idx"] or 0)) / 2.0
            else:  # after
                new_order = (rows[idx]["order_idx"] or 0) + 1.0 if idx == len(rows) - 1 \
                    else ((rows[idx]["order_idx"] or 0) + (rows[idx + 1]["order_idx"] or 0)) / 2.0
    cols = config.ALL_FIELDS
    ph = ", ".join(["?"] * len(cols))
    col_sql = ", ".join([f'"{c}"' for c in cols])
    defaults = default_values()
    vals = [defaults.get(c, "") for c in cols]
    conn.execute(f"INSERT INTO records (seq, order_idx, {col_sql}) VALUES (?, ?, {ph})",
                 [max_seq, new_order] + vals)
    new_id = conn.execute("SELECT last_insert_rowid() AS i").fetchone()["i"]
    bump_version(conn)
    conn.commit()
    conn.close()
    return jsonify({"ok": True, "id": new_id, "order_idx": new_order, "seq": max_seq})


@app.route("/api/row/<int:rid>", methods=["DELETE"])
def api_delete_row(rid):
    """软删除：只打标记进回收站，可在管理页恢复。"""
    user = request.args.get("user", "")
    conn = get_db()
    conn.execute(
        "UPDATE records SET is_deleted=1, deleted_at=?, deleted_by=? WHERE id=?",
        (datetime.now().strftime("%Y-%m-%d %H:%M:%S"), user, rid))
    bump_version(conn)
    conn.commit()
    conn.close()
    return jsonify({"ok": True})


# ====================== 回收站 ======================
@app.route("/api/trash")
def api_trash_list():
    conn = get_db()
    flds = ", ".join([f'"{f}"' for f in config.ALL_FIELDS])
    rows = conn.execute(
        f'SELECT id, seq, {flds}, deleted_at, deleted_by FROM records '
        f'WHERE COALESCE(is_deleted,0)=1 ORDER BY deleted_at DESC'
    ).fetchall()
    conn.close()
    return jsonify([dict(r) for r in rows])


@app.route("/api/trash/<int:rid>/restore", methods=["POST"])
def api_trash_restore(rid):
    conn = get_db()
    conn.execute("UPDATE records SET is_deleted=0, deleted_at=NULL, deleted_by=NULL WHERE id=?", (rid,))
    bump_version(conn)
    conn.commit()
    conn.close()
    return jsonify({"ok": True})


@app.route("/api/trash/<int:rid>", methods=["DELETE"])
def api_trash_purge(rid):
    """彻底删除（仅回收站里的记录）。"""
    conn = get_db()
    conn.execute("DELETE FROM records WHERE id=? AND COALESCE(is_deleted,0)=1", (rid,))
    bump_version(conn)
    conn.commit()
    conn.close()
    return jsonify({"ok": True})


def departure_year(conn, rid):
    """取某条记录当前的发班年份（用于锁定联动范围）。"""
    r = conn.execute('SELECT "发班时间" FROM records WHERE id=?', (rid,)).fetchone()
    dep = (r["发班时间"] or "").strip() if r else ""
    return dep[:4] if len(dep) >= 4 and dep[:4].isdigit() else ""


def sync_departure(conn, rid, field, value, now, user, old_year):
    """发班时间的班列级一致性：一个班列只有一个发班时间，改一条 → 同班列其他记录跟着改。

    年份锁定用【改动前的年份】(old_year)，这样把 2026 年的班列整体改到 2027 年时，
    同班列所有行仍能被一起带走；跨年同号的另一年班次不受影响。
    发班时间为空的行视为同批次待填，一并同步。目的站等其他字段不联动。
    返回被联动修改的行数（不含本行）。
    """
    if field != "发班时间" or not value or not str(value).strip():
        return 0
    r = conn.execute('SELECT "班列号" FROM records WHERE id=?', (rid,)).fetchone()
    train = (r["班列号"] or "").strip() if r else ""
    if not train:
        return 0
    year = old_year or str(value)[:4]
    cur = conn.execute(
        'UPDATE records SET "发班时间"=?, updated_at=?, updated_by=? '
        'WHERE "班列号"=? AND id<>? AND COALESCE(is_deleted,0)=0 '
        'AND "发班时间"<>? '
        'AND (substr("发班时间",1,4)=? OR "发班时间" IS NULL OR "发班时间"=\'\')',
        (value, now, user, train, rid, value, year))
    return cur.rowcount or 0


@app.route("/api/row/<int:rid>", methods=["PATCH"])
def api_update(rid):
    data = request.get_json(force=True, silent=True) or {}
    field = data.get("field")
    value = data.get("value", "")
    user = data.get("user", "")
    if field not in config.ALL_FIELDS:
        return jsonify({"ok": False, "msg": "非法字段"}), 400
    conn = get_db()
    now = datetime.now().strftime("%Y-%m-%d %H:%M")
    # 改「发班时间」前先取原值年份：联动范围必须按【原年份】锁定，
    # 否则把 2026 年的班列改成 2027 年日期时，同班列其他行会因年份不匹配而漏改。
    old_year = departure_year(conn, rid) if field == "发班时间" else ""
    # 写入侧归一：所有日期字段统一 YYYY-MM-DD，防止网页编辑又脏回去（体检C 1.2）
    if field in DATE_FIELDS:
        value = _norm_date(value)
    conn.execute(f'UPDATE records SET "{field}"=?, updated_at=?, updated_by=? WHERE id=?',
                 (value, now, user, rid))
    # 发班时间具有班列级一致性：同一班列号下的所有记录应共享同一个发班时间
    synced = sync_departure(conn, rid, field, value, now, user, old_year)
    bump_version(conn)
    conn.commit()
    conn.close()
    return jsonify({"ok": True, "synced": synced})


@app.route("/api/cells", methods=["POST"])
def api_cells():
    """批量保存（兜底用）：前端在页面关闭/刷新前用 sendBeacon 把未提交的单元格一次性发来。
    也支持普通调用。edits: [{id, field, value}]。"""
    data = request.get_json(force=True, silent=True) or {}
    edits = data.get("edits") or []
    user = data.get("user", "")
    if not isinstance(edits, list):
        return jsonify({"ok": False, "msg": "edits 需为数组"}), 400
    now = datetime.now().strftime("%Y-%m-%d %H:%M")
    conn = get_db()
    saved = 0
    for e in edits:
        field = e.get("field")
        rid = e.get("id")
        if field not in config.ALL_FIELDS or not isinstance(rid, int):
            continue
        value = e.get("value", "")
        old_year = departure_year(conn, rid) if field == "发班时间" else ""
        # 写入侧归一：所有日期字段统一 YYYY-MM-DD（体检C 1.2）
        if field in DATE_FIELDS:
            value = _norm_date(value)
        conn.execute(f'UPDATE records SET "{field}"=?, updated_at=?, updated_by=? WHERE id=?',
                     (value, now, user, rid))
        saved += 1
        # 兜底批量保存也保持发班时间的班列级一致性（与单格 PATCH 同一套规则）
        sync_departure(conn, rid, field, value, now, user, old_year)
    if saved:
        bump_version(conn)
    conn.commit()
    conn.close()
    return jsonify({"ok": True, "saved": saved})


# ====================== 时间戳回写接口（B方案P1：飞书→yxo.db 切换） ======================
# 仅供本机 DSK/ATB 转发机器人调用，按「箱号」匹配最新未删除记录，写 dsk/ATB 列。
# 带 token 校验，避免经 nginx 公网暴露被滥用。
STAMP_TOKEN = "yxo_stamp_local_2026"

@app.route("/api/stamp", methods=["POST"])
def api_stamp():
    if request.headers.get("X-Stamp-Token") != STAMP_TOKEN:
        return jsonify(ok=False, error="unauthorized"), 403
    data = request.get_json(force=True, silent=True) or {}
    box_no = (data.get("box_no") or "").strip()
    field = data.get("field")
    value = data.get("value") or ""
    if not box_no or field not in ("dsk", "ATB"):
        return jsonify(ok=False, error="bad params"), 400
    conn = get_db()
    try:
        # SQLite 不支持 UPDATE ... ORDER BY/LIMIT，先取最新未删除记录的 id 再按 id 更新
        row = conn.execute(
            'SELECT id FROM records WHERE "箱号"=? AND (is_deleted=0 OR is_deleted IS NULL) '
            'ORDER BY id DESC LIMIT 1', (box_no,)).fetchone()
        if not row:
            conn.close()
            return jsonify(ok=True, updated=0, found=False)
        now = datetime.now().strftime("%Y-%m-%d %H:%M")
        by = "DSK转发" if field == "dsk" else "ATB转发"
        conn.execute(f'UPDATE records SET "{field}"=?, updated_at=?, updated_by=? WHERE id=?',
                     (value, now, by, row["id"]))
        bump_version(conn)
        conn.commit()
        return jsonify(ok=True, updated=1, found=True)
    except Exception as e:
        try:
            conn.rollback()
        except Exception:
            pass
        conn.close()
        return jsonify(ok=False, error=str(e)), 500


@app.route("/api/price/<int:rid>", methods=["POST"])
def api_price(rid):
    conn = get_db()
    r = conn.execute(
        f'SELECT {", ".join(["\""+f+"\"" for f in config.ALL_FIELDS])} FROM records WHERE id=?',
        (rid,)).fetchone()
    conn.close()
    if not r:
        return jsonify({"ok": False}), 404
    price = compute_price(r["口岸"], r["目的站"], r["箱属"], r["发班时间"], load_price_config())
    if price is None:
        return jsonify({"ok": False, "msg": "无匹配价格（检查 口岸/目的站/发班时间）"})
    conn = get_db()
    conn.execute('UPDATE records SET "单箱价格"=?, updated_at=?, updated_by=? WHERE id=?',
                 (price, datetime.now().strftime("%Y-%m-%d %H:%M"), "算价", rid))
    bump_version(conn)
    conn.commit()
    conn.close()
    return jsonify({"ok": True, "price": price})


@app.route("/api/price/batch", methods=["POST"])
def api_price_batch():
    """对所有“单箱价格为空”的记录批量算价（已填价格的不覆盖）。"""
    conn = get_db()
    cfg = load_price_config()
    rows = conn.execute(
        f'SELECT id, {", ".join(["\""+f+"\"" for f in config.ALL_FIELDS])} FROM records').fetchall()
    priced = 0
    for r in rows:
        raw = r["单箱价格"]
        if raw is not None and str(raw).strip() not in ("", "0", "0.0", "0.00"):
            continue
        p = compute_price(r["口岸"], r["目的站"], r["箱属"], r["发班时间"], cfg)
        if p is None:
            continue
        conn.execute('UPDATE records SET "单箱价格"=?, updated_at=?, updated_by=? WHERE id=?',
                     (p, datetime.now().strftime("%Y-%m-%d %H:%M"), "批量算价", r["id"]))
        priced += 1
    bump_version(conn)
    conn.commit()
    conn.close()
    return jsonify({"ok": True, "priced": priced})


@app.route("/api/export", methods=["POST"])
def api_export():
    """按前端传入的“字段顺序 + 当前筛选行”导出 Excel（WPS 在线表风格：彩色表头/边框/冻结首行/自适应列宽）"""
    data = request.get_json(force=True, silent=True) or {}
    columns = data.get("columns") or config.ALL_FIELDS
    rows = data.get("rows") or []
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "订舱数据"
        thin = Side(style="thin", color="D0D5DD")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)
        head_fill = PatternFill("solid", fgColor="0F6E56")
        head_font = Font(bold=True, color="FFFFFF", size=11)
        ws.append(columns)
        for c in range(1, len(columns) + 1):
            cell = ws.cell(row=1, column=c)
            cell.fill = head_fill
            cell.font = head_font
            cell.border = border
            cell.alignment = Alignment(horizontal="center", vertical="center")
        for r in rows:
            ws.append([(r.get(col) if r.get(col) is not None else "") for col in columns])
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, max_col=len(columns)):
            for cell in row:
                cell.border = border
        for col in range(1, len(columns) + 1):
            cl = ws.cell(row=1, column=col).column_letter
            maxlen = len(str(columns[col - 1]))
            for rr in range(2, ws.max_row + 1):
                v = ws.cell(row=rr, column=col).value
                if v is not None:
                    maxlen = max(maxlen, len(str(v)))
            ws.column_dimensions[cl].width = min(max(maxlen + 2, 8), 40)
        ws.freeze_panes = "A2"
        out = os.path.join(config.DATA_DIR, "export_" + datetime.now().strftime("%Y%m%d%H%M%S") + ".xlsx")
        wb.save(out)
        return send_file(out, as_attachment=True, download_name="订舱数据_导出.xlsx")
    except Exception as e:
        return jsonify({"ok": False, "msg": "导出失败: " + str(e)}), 500


# ====================== 托书自动生成 ======================
# 引擎由芙蕾雅提供（tuoshu_engine.py，纯逻辑无 Flask 依赖），这里只做集成。
def _check_tuoshu_user():
    """托书权限：仅 config.TUOSHU_ADMINS（独立分组，不复用系统管理的 LIMITED_ADMINS）"""
    u = request.args.get("user") or (request.get_json(silent=True) or {}).get("user", "")
    return u in config.TUOSHU_ADMINS


def _tuoshu_forbid():
    return jsonify({"ok": False, "msg": "无权限：托书生成仅限指定人员使用"}), 403


# 所有"日期"类型字段（config.FIELD_DEFS 里 type=="date"）：发班时间 / 入堆场 / 入站
# 写入时统一归一成 YYYY-MM-DD，防止网页/导入又脏回斜杠、点、反斜杠等分隔符。
DATE_FIELDS = {f["name"] for f in config.FIELD_DEFS if f.get("type") == "date"}


def _norm_date(s):
    """日期归一成 YYYY-MM-DD（库现行标准，横杠）。

    历史数据/网页旧逻辑曾用 2026/08/29 斜杠格式，引擎按 %Y-%m-%d 解析会抛异常，
    故统一归一。这里同时兼容 / . \\ 三种分隔符，归一后只保留横杠。"""
    s = str(s or "").strip()[:10]
    if not s:
        return ""
    s = s.replace("/", "-").replace(".", "-").replace("\\", "-")
    parts = s.split("-")
    if len(parts) != 3:
        return ""
    try:
        return "%04d-%02d-%02d" % (int(parts[0]), int(parts[1]), int(parts[2]))
    except ValueError:
        return ""


def _load_dest_map(conn):
    """读 tuoshu_dest_map → (dest_map, dest_normalize)，注入引擎替代硬编码字典"""
    dest_map, dest_norm = {}, {}
    try:
        cur = conn.execute("SELECT station_cn, en_name, country, alias FROM tuoshu_dest_map")
    except sqlite3.OperationalError:
        return dest_map, dest_norm
    for r in cur:
        station = (r["station_cn"] or "").strip()
        if not station:
            continue
        en = (r["en_name"] or "").strip()
        country = (r["country"] or "").strip()
        dest_map[station] = ("%s, %s" % (en, country)) if country else en
        for a in (r["alias"] or "").split(","):
            a = a.strip()
            if a:
                dest_norm[a] = station
    return dest_map, dest_norm


def _tuoshu_groups(conn, data):
    """按筛选条件查 records 并聚合成托书分组。

    聚合维度 group_by：
      train_dest（默认）= (班列号, 发班年份, 目的站)。同一班列常挂多个目的站的箱子，
                          而托书 E15 只能写一个目的站，按班列合并会把别家的箱量算进来。
      train              = (班列号, 发班年份)，原引擎口径，目的站取组内首个。
    与 Dashboard 一致的是「按发班年份锁定」（班列号每年重置，跨年同名≠同班）。
    箱量 = 有效记录行数：托书在「订舱阶段」生成，此时箱号普遍尚未录入，
           按箱号计数会恒为 0，所以按行数计。
    """
    date_from = _norm_date(data.get("date_from"))
    date_to = _norm_date(data.get("date_to"))
    train_nos = [str(t).strip() for t in (data.get("train_nos") or []) if str(t).strip()]
    train_type = (data.get("train_type") or "").strip()
    group_by = (data.get("group_by") or "train").strip()

    sql = ('SELECT "班列号","发班时间","目的站","箱号","班列类型","状态" FROM records '
           "WHERE COALESCE(is_deleted,0)=0 AND IFNULL(\"班列号\",'')<>''")
    params = []
    if train_type:
        sql += ' AND "班列类型"=?'
        params.append(train_type)
    if train_nos:
        sql += ' AND "班列号" IN (%s)' % ",".join("?" * len(train_nos))
        params.extend(train_nos)

    dest_map, dest_norm = _load_dest_map(conn)
    groups = {}
    for r in conn.execute(sql, params):
        dep = _norm_date(r["发班时间"])
        if not dep:
            continue
        if date_from and dep < date_from:
            continue
        if date_to and dep > date_to:
            continue
        if (r["状态"] or "").strip() == "退舱":   # 退舱不计入托书箱量
            continue
        tno = (r["班列号"] or "").strip()
        st = (r["目的站"] or "").strip()
        key = (tno, dep, st) if group_by == "train_dest" else (tno, dep)
        g = groups.get(key)
        if not g:
            g = groups[key] = {
                "train_no": tno,
                "departure_date": dep,
                "station_cn": st,
                "train_type": (r["班列类型"] or "").strip(),
                "boxes": 0,
                "_stations": set(),
            }
        g["boxes"] += 1
        if st:
            g["_stations"].add(st)
            if not g["station_cn"]:
                g["station_cn"] = st

    out = []
    for g in groups.values():
        stations = sorted(g.pop("_stations"))
        # 目的站可选项 + 各自解析结果（前端下拉用）
        dest_options = {}
        for s in stations:
            n = dest_norm.get(s, s)
            dest_options[s] = dest_map.get(n, "") or ("未知目的站(%s)" % s if s else "未知目的站")
        s = g["station_cn"]
        resolved = bool(dest_options.get(s)) and not str(dest_options.get(s, "")).startswith("未知目的站")
        g["station_options"] = stations
        g["dest_options"] = dest_options
        g["destination"] = dest_options.get(s, "未知目的站")
        g["resolved"] = resolved
        warns = []
        if not resolved:
            warns.append("目的站「%s」未配置英文映射" % (s or "(空)"))
        if len(stations) > 1:
            warns.append("同班列同日期存在多个目的站：%s，请在列表中选定" % "/".join(stations))
        g["warn"] = "；".join(warns)
        out.append(g)
    out.sort(key=lambda x: (x["departure_date"], x["train_no"]))
    return out, dest_map, dest_norm


def _load_template(conn, template_id=None):
    """取模板登记行 → (file_path, field_map, name_pattern)"""
    import tuoshu_engine
    row = None
    try:
        if template_id:
            row = conn.execute(
                "SELECT * FROM tuoshu_templates WHERE id=?", (template_id,)).fetchone()
        if not row:
            row = conn.execute(
                "SELECT * FROM tuoshu_templates ORDER BY id LIMIT 1").fetchone()
    except sqlite3.OperationalError:
        row = None
    if not row:
        return (os.path.join(config.DATA_DIR, "tuoshu_templates", "托书模板.xlsx"),
                tuoshu_engine.DEFAULT_FIELD_MAP, tuoshu_engine.DEFAULT_NAME_PATTERN)
    try:
        fm_raw = json.loads(row["field_map_json"] or "{}")
        field_map = {k: tuple(v) for k, v in fm_raw.items()} or tuoshu_engine.DEFAULT_FIELD_MAP
    except Exception:
        field_map = tuoshu_engine.DEFAULT_FIELD_MAP
    return (row["file_path"], field_map,
            row["name_pattern"] or tuoshu_engine.DEFAULT_NAME_PATTERN)


@app.route("/tuoshu")
def tuoshu_page():
    from flask import make_response
    resp = make_response(render_template("tuoshu.html"))
    resp.headers["Cache-Control"] = "no-store, no-cache, must-revalidate, max-age=0"
    resp.headers["Pragma"] = "no-cache"
    resp.headers["Expires"] = "0"
    return resp


@app.route("/api/tuoshu/meta")
def api_tuoshu_meta():
    """页面初始化：模板列表 + 可选班列号（近 6 个月）"""
    if not _check_tuoshu_user():
        return _tuoshu_forbid()
    conn = get_db()
    templates = []
    try:
        for r in conn.execute("SELECT id,name FROM tuoshu_templates ORDER BY id"):
            templates.append({"id": r["id"], "name": r["name"]})
    except sqlite3.OperationalError:
        pass
    conn.close()
    return jsonify({"ok": True, "templates": templates,
                    "train_types": ["散舱", "专列"]})


@app.route("/api/tuoshu/preview", methods=["POST"])
def api_tuoshu_preview():
    """生成前预检：列出将要生成的托书、箱量与目的站解析结果（未映射会明确告警）"""
    if not _check_tuoshu_user():
        return _tuoshu_forbid()
    data = request.get_json(force=True, silent=True) or {}
    conn = get_db()
    try:
        groups, _, _ = _tuoshu_groups(conn, data)
    except Exception as e:
        conn.close()
        return jsonify({"ok": False, "msg": "预检失败: " + str(e)}), 500
    conn.close()
    return jsonify({"ok": True, "groups": groups,
                    "total": len(groups),
                    "boxes": sum(g["boxes"] for g in groups),
                    "unresolved": [g["train_no"] for g in groups if not g["resolved"]]})


@app.route("/api/tuoshu/generate", methods=["POST"])
def api_tuoshu_generate():
    """按筛选条件批量生成托书 xlsx，多于一份时打 zip 返回"""
    if not _check_tuoshu_user():
        return _tuoshu_forbid()
    import shutil
    import zipfile
    import tuoshu_engine

    data = request.get_json(force=True, silent=True) or {}
    booking_date = _norm_date(data.get("booking_date")) or datetime.now().strftime("%Y-%m-%d")

    conn = get_db()
    try:
        groups, dest_map, dest_norm = _tuoshu_groups(conn, data)
        tpl_path, field_map, name_pattern = _load_template(conn, data.get("template_id"))
    except Exception as e:
        conn.close()
        return jsonify({"ok": False, "msg": "查询失败: " + str(e)}), 500
    conn.close()

    # 前端预检后可能只勾选了部分分组。
    # train_dest：按 (班列号,发班日期,目的站) 精确匹配；
    # train（按班列合并）：按 (班列号,发班日期) 匹配，目的站用用户勾选时选定的。
    picked = data.get("picked") or []
    if picked:
        group_by = (data.get("group_by") or "train").strip()
        if group_by == "train":
            sel = {}
            for p in picked:
                sel[(str(p.get("train_no") or ""), str(p.get("departure_date") or ""))] = \
                    str(p.get("station_cn") or "")
            new_groups = []
            for g in groups:
                st = sel.get((g["train_no"], g["departure_date"]))
                if st is None:
                    continue
                g = dict(g)
                g["station_cn"] = st   # 用用户选定的目的站生成
                new_groups.append(g)
            groups = new_groups
        else:
            keys = {(str(p.get("train_no") or ""), str(p.get("departure_date") or ""),
                     str(p.get("station_cn") or "")) for p in picked}
            groups = [g for g in groups
                      if (g["train_no"], g["departure_date"], g["station_cn"]) in keys]

    if not groups:
        return jsonify({"ok": False, "msg": "所选条件下没有可生成的班列"}), 400
    if not os.path.exists(tpl_path):
        return jsonify({"ok": False, "msg": "模板文件不存在: " + tpl_path}), 500

    stamp = datetime.now().strftime("%Y%m%d%H%M%S")
    root_out = os.path.join(config.DATA_DIR, "tuoshu_out")
    out_dir = os.path.join(root_out, stamp)
    os.makedirs(out_dir, exist_ok=True)
    _purge_old_tuoshu(root_out, keep_days=7)   # 清理历史产物，避免无限堆积

    try:
        eng = tuoshu_engine.TuoshuEngine(
            tpl_path, field_map=field_map, name_pattern=name_pattern,
            dest_map=dest_map, dest_normalize=dest_norm)
        files = []
        for g in groups:
            files.append(eng.generate_one(
                {"train_no": g["train_no"], "departure_date": g["departure_date"],
                 "station_cn": g["station_cn"], "boxes": g["boxes"]},
                booking_date, out_dir))

        # 注意：send_file 是惰性发送，这里不能删除正在返回的文件所在目录。
        if len(files) == 1:
            return send_file(files[0], as_attachment=True,
                             download_name=os.path.basename(files[0]))
        zip_path = os.path.join(config.DATA_DIR, "托书_%s.zip" % stamp)
        with zipfile.ZipFile(zip_path, "w", zipfile.ZIP_DEFLATED) as zf:
            for f in files:
                zf.write(f, os.path.basename(f))
        shutil.rmtree(out_dir, ignore_errors=True)   # zip 已含全部内容，可安全删源目录
        return send_file(zip_path, as_attachment=True,
                         download_name="渝新欧订舱委托书_%s.zip" % stamp)
    except Exception as e:
        shutil.rmtree(out_dir, ignore_errors=True)
        return jsonify({"ok": False, "msg": "生成失败: " + str(e)}), 500


@app.route("/api/tuoshu/dest_map", methods=["GET"])
def api_tuoshu_dest_map_get():
    """列出目的站→英文映射，并提示 records 里已出现但未配置映射的目的站"""
    if not _check_tuoshu_user():
        return _tuoshu_forbid()
    conn = get_db()
    try:
        rows = []
        for r in conn.execute(
                "SELECT station_cn, en_name, country, alias FROM tuoshu_dest_map ORDER BY station_cn"):
            rows.append({"station_cn": r["station_cn"], "en_name": r["en_name"],
                         "country": r["country"], "alias": r["alias"]})
        used = set()
        try:
            for r in conn.execute(
                    'SELECT DISTINCT "目的站" FROM records '
                    'WHERE COALESCE(is_deleted,0)=0 AND TRIM(COALESCE("目的站",""))<>""'):
                s = (r["目的站"] or "").strip()
                if s:
                    used.add(s)
        except sqlite3.OperationalError:
            pass
        mapped = {row["station_cn"] for row in rows}
        used_unmapped = sorted(used - mapped)
        return jsonify({"ok": True, "rows": rows, "used_unmapped": used_unmapped})
    except Exception as e:
        return jsonify({"ok": False, "msg": "加载失败: " + str(e)}), 500
    finally:
        conn.close()


@app.route("/api/tuoshu/dest_map", methods=["POST"])
def api_tuoshu_dest_map_post():
    """新增 / 编辑 / 删除 目的站→英文映射（tuoshu_dest_map 表，主键 station_cn）"""
    if not _check_tuoshu_user():
        return _tuoshu_forbid()
    d = request.get_json(force=True, silent=True) or {}
    station = (d.get("station_cn") or "").strip()
    if not station:
        return jsonify({"ok": False, "msg": "目的站(中文)不能为空"}), 400
    conn = get_db()
    try:
        if d.get("_delete"):
            conn.execute("DELETE FROM tuoshu_dest_map WHERE station_cn=?", (station,))
            conn.commit()
            return jsonify({"ok": True, "msg": "已删除"})
        conn.execute(
            """INSERT INTO tuoshu_dest_map(station_cn, en_name, country, alias)
               VALUES(?,?,?,?)
               ON CONFLICT(station_cn) DO UPDATE SET
                 en_name=excluded.en_name, country=excluded.country, alias=excluded.alias""",
            (station, (d.get("en_name") or "").strip(), (d.get("country") or "").strip(),
             (d.get("alias") or "").strip()))
        conn.commit()
        return jsonify({"ok": True, "msg": "已保存"})
    except Exception as e:
        conn.rollback()
        return jsonify({"ok": False, "msg": "保存失败: " + str(e)}), 500
    finally:
        conn.close()


def _purge_old_tuoshu(root_out, keep_days=7):
    """清理 N 天前的托书产物目录与 zip，避免 data/ 无限膨胀"""
    import shutil
    import time
    cutoff = time.time() - keep_days * 86400
    try:
        for name in os.listdir(root_out):
            p = os.path.join(root_out, name)
            if os.path.isdir(p) and os.path.getmtime(p) < cutoff:
                shutil.rmtree(p, ignore_errors=True)
    except Exception:
        pass
    try:
        for name in os.listdir(config.DATA_DIR):
            if name.startswith("托书_") and name.endswith(".zip"):
                p = os.path.join(config.DATA_DIR, name)
                if os.path.getmtime(p) < cutoff:
                    os.remove(p)
    except Exception:
        pass


@app.route("/api/import", methods=["POST"])
def api_import():
    conn = get_db()
    n = run_import(conn)
    bump_version(conn)
    conn.commit()
    conn.close()
    return jsonify({"ok": True, "imported": n})


@app.route("/api/import_upload", methods=["POST"])
def api_import_upload():
    """接收用户上传的 Excel 文件并导入。
    参数：train_type=散舱|专列（必带）；force=1 跳过预判确认。
    预判：同班列票数 >=30 视作疑似专列、<15 视作疑似散舱，与所选类型矛盾时
    返回 need_confirm 让前端二次确认，避免整表标错难以修正。"""
    from import_excel import analyze_import
    f = request.files.get("file")
    if not f:
        return jsonify({"ok": False, "msg": "未收到文件"}), 400
    train_type = (request.form.get("train_type") or "散舱").strip()
    if train_type not in ("散舱", "专列"):
        train_type = "散舱"
    force = request.form.get("force") == "1"
    suffix = os.path.splitext(f.filename or "x.xlsx")[1] or ".xlsx"
    tmp = os.path.join(config.DATA_DIR, "_upload_" + datetime.now().strftime("%Y%m%d%H%M%S") + suffix)
    try:
        f.save(tmp)
        # —— 散舱/专列预判 ——
        if not force:
            counts = analyze_import(tmp)
            if counts:
                big = {t: n for t, n in counts.items() if n >= 30}
                small_max = max(counts.values())
                warn = None
                if train_type == "散舱" and big:
                    desc = "、".join(f"{t}（{n} 票）" for t, n in big.items())
                    warn = f"检测到 {desc} 同班列票数很多，更像【专列】。确定按【散舱】导入吗？"
                elif train_type == "专列" and small_max < 15:
                    warn = f"本表最大同班列票数只有 {small_max} 票，更像【散舱】。确定按【专列】导入吗？"
                if warn:
                    return jsonify({"ok": False, "need_confirm": True, "msg": warn})
        conn = get_db()
        n = run_import(conn, tmp, train_type=train_type)
        bump_version(conn)
        conn.commit()
        conn.close()
    except Exception as e:
        return jsonify({"ok": False, "msg": "导入失败: " + str(e)}), 500
    finally:
        try:
            os.remove(tmp)
        except Exception:
            pass
    return jsonify({"ok": True, "imported": n, "train_type": train_type})


# ====================== 舱单 / 箱号 统一导入（#199）======================
# 引擎 manifest_engine.py 由小叽实现（纯逻辑，无 Flask 依赖），这里只做集成 + 权限闸门。
def _check_manifest_user():
    """舱单导入权限：仅 config.MANIFEST_ADMIN（毛骁洋）。"""
    u = request.args.get("user") or (request.get_json(silent=True) or {}).get("user", "")
    return u == config.MANIFEST_ADMIN


def _manifest_forbid():
    return jsonify({"ok": False, "msg": "无权限：舱单导入仅限 %s 使用" % config.MANIFEST_ADMIN}), 403


@app.route("/manifest")
def manifest_page():
    from flask import make_response
    resp = make_response(render_template("manifest.html"))
    resp.headers["Cache-Control"] = "no-store, no-cache, must-revalidate, max-age=0"
    resp.headers["Pragma"] = "no-cache"
    resp.headers["Expires"] = "0"
    return resp


@app.route("/api/manifest/upload", methods=["POST"])
def api_manifest_upload():
    """空跑：解析上传的 xlsx → 差异清单，不写库。"""
    if not _check_manifest_user():
        return _manifest_forbid()
    f = request.files.get("file")
    if not f:
        return jsonify({"ok": False, "msg": "未收到文件"}), 400
    import io
    try:
        from manifest_engine import parse_workbook_diagnostics, normalize_row, build_diff
        data = f.read()
        headers, rows, ftype, diag = parse_workbook_diagnostics(io.BytesIO(data))
        parsed = [normalize_row(r, ftype) for r in rows]
        conn = get_db()
        diff = build_diff(conn, parsed)
        conn.close()
        resp = {
            "ok": True, "ftype": ftype, "row_count": len(parsed),
            "headers": [str(h) for h in headers], "diff": diff,
            "diagnostics": diag,
        }
        # 当结果异常时，确保前端能看到诊断信息
        if ftype == "unknown" or len(parsed) == 0:
            resp["diagnostics"]["error_summary"] = (
                diag.get("error") or
                (f"文件已读取但未能识别为舱单/箱号模板。"
                 f"表头行第{diag.get('header_row_index','?')}行，"
                 f"共{diag.get('data_stats',{}).get('total_data_lines','?')}行数据，"
                 f"解析有效行{diag.get('data_stats',{}).get('parsed_rows',0)}。"
                 f"请检查：(1)文件是否为xlsx格式(非xls) (2)首行是否包含「客户编码」列 "
                 f"(3)是否有多余的标题行遮挡表头")
            )
        return jsonify(resp)
    except Exception as e:
        import traceback
        return jsonify({"ok": False, "msg": "解析失败: " + str(e),
                        "traceback": traceback.format_exc()}), 500


@app.route("/api/manifest/apply", methods=["POST"])
def api_manifest_apply():
    """应用：快照 → 事务写入 → 留痕。失败整体回滚。"""
    if not _check_manifest_user():
        return _manifest_forbid()
    body = request.get_json(silent=True) or {}
    diff = body.get("diff")
    if not diff:
        return jsonify({"ok": False, "msg": "缺少 diff"}), 400
    operator = body.get("user", config.MANIFEST_ADMIN)
    files = body.get("files", [])
    if isinstance(files, str):
        files = [files]
    from manifest_engine import apply_diff
    conn = get_db()
    try:
        batch_id = apply_diff(conn, diff, operator, files)
        conn.commit()
    except Exception as e:
        try:
            conn.rollback()
        except Exception:
            pass
        return jsonify({"ok": False, "msg": "应用失败(已回滚): " + str(e)}), 500
    finally:
        conn.close()
    return jsonify({"ok": True, "batch_id": batch_id})


@app.route("/api/manifest/batches", methods=["GET"])
def api_manifest_batches():
    if not _check_manifest_user():
        return _manifest_forbid()
    conn = get_db()
    rows = conn.execute(
        "SELECT batch_id,batch_type,source_files,COALESCE(n_update,0) n_update,"
        "COALESCE(n_insert,0) n_insert,COALESCE(n_alert,0) n_alert,operator,"
        "COALESCE(reverted,0) reverted,created_at FROM import_batch "
        "ORDER BY created_at DESC LIMIT 100"
    ).fetchall()
    conn.close()
    return jsonify({"ok": True, "batches": [dict(r) for r in rows]})


@app.route("/api/manifest/batch/<batch_id>", methods=["GET"])
def api_manifest_batch_detail(batch_id):
    if not _check_manifest_user():
        return _manifest_forbid()
    conn = get_db()
    logs = conn.execute(
        "SELECT id,batch_id,batch_type,record_id,\"客户编码\",\"箱号\",field,"
        "old_value,new_value,action,source_file,operator,COALESCE(reverted,0) reverted,"
        "created_at FROM update_log WHERE batch_id=? ORDER BY id",
        (batch_id,)
    ).fetchall()
    conn.close()
    return jsonify({"ok": True, "logs": [dict(r) for r in logs]})


@app.route("/api/manifest/revert", methods=["POST"])
def api_manifest_revert():
    """整批回退（batch_id）或单条撤销（log_id）。"""
    if not _check_manifest_user():
        return _manifest_forbid()
    body = request.get_json(silent=True) or {}
    from manifest_engine import revert_batch, revert_item
    conn = get_db()
    try:
        if body.get("log_id"):
            ok = revert_item(conn, body["log_id"])
            if not ok:
                return jsonify({"ok": False, "msg": "该条已回退或不存在"}), 400
        elif body.get("batch_id"):
            revert_batch(conn, body["batch_id"])
        else:
            return jsonify({"ok": False, "msg": "缺少 batch_id 或 log_id"}), 400
        conn.commit()
    except Exception as e:
        try:
            conn.rollback()
        except Exception:
            pass
        return jsonify({"ok": False, "msg": "回退失败: " + str(e)}), 500
    finally:
        conn.close()
    return jsonify({"ok": True})


@app.route("/api/manifest/restore", methods=["POST"])
def api_manifest_restore():
    """核弹级：用批次快照整库还原。前端需二次确认。"""
    if not _check_manifest_user():
        return _manifest_forbid()
    body = request.get_json(silent=True) or {}
    batch_id = body.get("batch_id")
    if not batch_id:
        return jsonify({"ok": False, "msg": "缺少 batch_id"}), 400
    from manifest_engine import restore_snapshot
    conn = get_db()
    try:
        snap = restore_snapshot(conn, batch_id)
    except Exception as e:
        try:
            conn.close()
        except Exception:
            pass
        return jsonify({"ok": False, "msg": "还原失败: " + str(e)}), 500
    conn.close()
    return jsonify({"ok": True, "snapshot": snap})


# —— 个人视图状态（按用户隔离，互不干扰）——
@app.route("/api/state", methods=["GET", "POST"])
def api_state():
    user = request.args.get("user") or (request.get_json(silent=True) or {}).get("user", "")
    if not user:
        return jsonify({"ok": False, "msg": "缺少 user"}), 400
    conn = get_db()
    if request.method == "POST":
        payload = request.get_json(silent=True) or {}
        val = json.dumps(payload.get("state", {}), ensure_ascii=False)
        conn.execute(
            "INSERT INTO user_state (user, key, value) VALUES (?, 'state', ?) "
            "ON CONFLICT(user, key) DO UPDATE SET value=excluded.value",
            (user, val))
        conn.commit()
        conn.close()
        return jsonify({"ok": True})
    row = conn.execute("SELECT value FROM user_state WHERE user=? AND key='state'", (user,)).fetchone()
    conn.close()
    if row:
        try:
            return jsonify({"ok": True, "state": json.loads(row["value"])})
        except Exception:
            return jsonify({"ok": True, "state": {}})
    return jsonify({"ok": True, "state": {}})


def init_user_table():
    conn = get_db()
    conn.execute("""
        CREATE TABLE IF NOT EXISTS user_state (
            user TEXT, key TEXT, value TEXT,
            PRIMARY KEY (user, key)
        )
    """)
    conn.execute("""
        CREATE TABLE IF NOT EXISTS meta_kv (
            key TEXT PRIMARY KEY, value TEXT
        )
    """)
    conn.commit()
    conn.close()


if __name__ == "__main__":
    init_user_table()
    init_db()
    print(f"YXO 订舱数据管理已启动: http://0.0.0.0:{config.PORT}  (反代路径 /yxo/)")
    app.run(host="0.0.0.0", port=config.PORT, debug=False, threaded=True)
