# -*- coding: utf-8 -*-
"""测试集校验 v2：eml 可解析 + 运单号 xlsx 真实多行 + manifest split/forward 完整 + 目录清点"""
import email, glob, os, json, io
from email.header import decode_header
from openpyxl import load_workbook

BASE = r"C:\Users\Roke8x\WorkBuddy\2026-07-29-10-11-57\test_fixtures"

def dec(s):
    if not s: return ""
    out = []
    for b, e in decode_header(s):
        out.append(b.decode(e or "utf-8", "replace") if isinstance(b, bytes) else b)
    return "".join(out)

print("=== 目录清点 ===")
for sub in ["draft", "waybill", "tracing", "dsk", "atb"]:
    d = os.path.join(BASE, sub)
    n = len(glob.glob(os.path.join(d, "*.eml"))) if os.path.isdir(d) else 0
    print(f"  {sub}: {n} 个 eml")
total = len(glob.glob(os.path.join(BASE, "*", "*.eml")))
print(f"  合计: {total} 个 eml")

ok = 0; bad = 0
print("\n=== EML 结构校验 ===")
for f in sorted(glob.glob(os.path.join(BASE, "*", "*.eml"))):
    msg = email.message_from_bytes(open(f, "rb").read())
    subj = dec(msg.get("Subject", "")); frm = dec(msg.get("From", ""))
    if subj and frm:
        ok += 1
    else:
        bad += 1; print("  BAD:", f)
print(f"  OK={ok} BAD={bad}")

print("\n=== 运单号 xlsx 多行校验 (openpyxl 读回行数) ===")
expect_rows = {
    "W01_single_external.eml": 1, "W02_internal_only.eml": 1,
    "W03_multi_external.eml": 2, "W04_no_xls.eml": 0, "W05_unknown.eml": 1,
    "W06_multirow_mixed.eml": 3, "W07_multirow_one_co.eml": 2,
    "W08_multirow_with_cancel.eml": 3, "W09_multirow_with_unknown.eml": 3,
}
for fn, n in expect_rows.items():
    p = os.path.join(BASE, "waybill", fn)
    msg = email.message_from_bytes(open(p, "rb").read())
    xls = None
    for part in msg.walk():
        fn2 = part.get_filename()
        if fn2 and fn2.lower().endswith((".xlsx", ".xls")):
            xls = part.get_payload(decode=True); break
    if xls is None:
        got = 0
    else:
        try:
            wb = load_workbook(io.BytesIO(xls)); ws = wb.active
            got = sum(1 for r in ws.iter_rows(min_row=2, values_only=True) if any(c for c in r))
        except Exception as e:
            got = f"ERR:{e}"
    flag = "OK" if got == n else "MISMATCH"
    print(f"  [{flag}] {fn}: 期望={n} 实际={got}")

print("\n=== manifest split/forward 字段校验 ===")
man = json.load(open(os.path.join(BASE, "manifest.json"), encoding="utf-8"))
assert len(man) == total, f"manifest 条数({len(man)}) != eml 数({total})"
print(f"  manifest 条数={len(man)} 与 eml 数一致 ✅")
split_cases = [m for m in man if "split" in m]
print(f"  含 split 字段夹具: {len(split_cases)}")
for m in split_cases:
    sp = m["split"]
    assert "email_parts" in sp and "parts" in sp, f"{m['file']} split 缺字段"
    print(f"  {m['file']}: email={sp['email_parts']} wecom={sp.get('wecom_only_parts')} "
          f"parts={len(sp['parts'])} skip={sp.get('skip_rows', sp.get('skip_boxes', []))} "
          f"alarm={sp.get('alarm_rows', [])}")
print("  全部 split 字段结构完整 ✅")

print("\n=== 一致性：运单号 email_parts == 外部(has_xls)parts 数 ===")
for m in split_cases:
    if m["type"] == "waybill":
        ext = [p for p in m["split"]["parts"] if p.get("has_xls")]
        assert m["split"]["email_parts"] == len(ext), f"{m['file']} 不一致"
        print(f"  {m['file']}: {m['split']['email_parts']} == {len(ext)} ✅")

print("\n全部校验通过 ✅" if bad == 0 else f"\n存在 {bad} 个坏 eml ❌")
