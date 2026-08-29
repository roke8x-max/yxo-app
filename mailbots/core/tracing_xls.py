# -*- coding: utf-8 -*-
"""运踪 .xls 解析器（spec 刀 6：新增 tracing .xls 箱号解析能力）

功能：解析运踪邮件中的 .xls 附件，提取箱号、目的地、节点等信息
用于扇形转发：按箱号匹配负责公司，实现扇形转发
"""
import xlrd
import openpyxl
import re
from typing import List, Dict, Optional, Tuple, Any
from io import BytesIO


def parse_tracing_xls(raw_bytes: bytes) -> List[Dict[str, Any]]:
    """
    解析运踪 .xls/.xlsx 附件，提取箱号信息
    
    返回格式：
    [
        {
            "train_no": "WB758",           # 班列号
            "container_no": "HLXU8152547",  # 箱号
            "container_type": "40HQ",       # 箱型
            "chinese_wagon": "中方车号",     # 中方车号
            "destination": "别雷拉斯特",     # 目的地
            "departure_time": "2026-08-16 02:30:00",  # 发车时间
            "node": "重庆",                  # 当前节点
            "status": "在途"                # 状态
        },
        ...
    ]
    """
    import io
    
    # 处理空/无效输入
    if not raw_bytes or len(raw_bytes) == 0:
        return []
    
    # 尝试不同的解析器
    for parser in [_parse_with_xlrd, _parse_with_openpyxl]:
        try:
            result = parser(raw_bytes)
            if result:
                return result
        except Exception as e:
            continue
    
    # 如果都失败，尝试文本解析
    return _parse_as_text(raw_bytes)

parse_tracing_xls_attachment = parse_tracing_xls


def _parse_with_xlrd(raw_bytes: bytes) -> List[Dict[str, Any]]:
    """使用 xlrd 解析 .xls 文件"""
    import xlrd
    
    try:
        workbook = xlrd.open_workbook(file_contents=raw_bytes)
    except Exception:
        # 无效的 xls 文件格式
        return []
        
    results = []
    
    for sheet in workbook.sheets():
        # 查找包含箱号的 sheet
        if sheet.nrows < 3 or sheet.ncols < 3:
            continue
            
        # 尝试识别表头行
        header_row = _find_header_row(sheet)
        if header_row < 0:
            continue
            
        # 解析列映射
        col_map = _map_columns(sheet, header_row)
        if not _has_required_columns(col_map):
            continue
            
        # 解析数据行
        for row_idx in range(header_row + 1, sheet.nrows):
            row_data = _parse_row(sheet, row_idx, col_map)
            if row_data:
                row_data['source'] = 'xlrd'
                results.append(row_data)
    
    return results


def _parse_with_openpyxl(raw_bytes: bytes) -> List[Dict[str, Any]]:
    """使用 openpyxl 解析 .xlsx 文件"""
    import openpyxl
    import io
    
    try:
        workbook = openpyxl.load_workbook(io.BytesIO(raw_bytes), read_only=True, data_only=True)
    except Exception:
        return []
        
    results = []
    
    for sheet in workbook.worksheets:
        if sheet.max_row < 3 or sheet.max_column < 3:
            continue
            
        # 查找表头行
        header_row = _find_header_row_openpyxl(sheet)
        if header_row < 0:
            continue
            
        col_map = _map_columns_openpyxl(sheet, header_row)
        if not _has_required_columns(col_map):
            continue
            
        for row_idx in range(header_row + 1, sheet.max_row + 1):
            row_data = _parse_row_openpyxl(sheet, row_idx, col_map)
            if row_data:
                row_data['source'] = 'openpyxl'
                results.append(row_data)
    
    return results


def _parse_as_text(raw_bytes: bytes) -> List[Dict[str, Any]]:
    """尝试将 xls 作为文本解析（针对 CSV 格式的 xls）"""
    import io
    text = raw_bytes.decode('utf-8', errors='ignore')
    
    lines = text.strip().split('\n')
    if len(lines) < 2:
        return []
    
    # 尝试 CSV 解析
    import csv
    reader = csv.reader(io.StringIO(text))
    rows = list(reader)
    
    if len(rows) < 2:
        return []
        
    header = [h.strip() for h in rows[0]]
    col_map = _map_columns_text(header)
    if not _has_required_columns(col_map):
        return []
        
    results = []
    for row in rows[1:]:
        row_data = _parse_row_text(row, col_map)
        if row_data:
            row_data['source'] = 'text'
            results.append(row_data)
    
    return results


# ============================================================
# 辅助函数
# ============================================================

def _find_header_row(sheet) -> int:
    """查找包含箱号相关列的表头行"""
    keywords = ['箱号', '箱号', 'container', 'container_no', 'container_no.', '箱型', '目的站', 'destination', '到站', '运单号', 'waybill']
    
    for row_idx in range(min(10, sheet.nrows)):
        row_values = [str(sheet.cell_value(row_idx, col)).strip().lower() 
                     for col in range(sheet.ncols)]
        if any(kw in ' '.join(row_values).lower() for kw in ['箱号', 'container', '箱型', '目的站', 'destination']):
            return row_idx
    return -1


def _find_header_row_openpyxl(sheet) -> int:
    keywords = ['箱号', 'container', 'container_no', '箱型', '目的站', 'destination', '到站', '运单号', 'waybill']
    
    for row_idx in range(1, min(11, sheet.max_row + 1)):
        row_values = [str(sheet.cell(row=row_idx, column=col).value or '').strip().lower() 
                     for col in range(1, sheet.max_column + 1)]
        if any(kw in ' '.join(row_values).lower() for kw in ['箱号', 'container', '箱型', '目的站', 'destination']):
            return row_idx
    return -1


def _map_columns(sheet, header_row: int) -> Dict[str, int]:
    """映射列名到列索引"""
    keywords = {
        'container_no': ['箱号', 'container', 'container_no', 'container_no.', '箱号'],
        'container_type': ['箱型', 'container_type', '类型'],
        'chinese_wagon': ['中方车号', '中方车号', 'chinese_wagon'],
        'container_no_alt': ['箱号', 'container_no', 'container_no.'],
        'destination': ['目的站', 'destination', '到站', '到站'],
        'chinese_wagon_no': ['中方车号', '中方车号', 'chinese_wagon'],
        'train_no': ['班列号', 'train_no', 'train_no', '班次'],
        'departure_time': ['发车时间', 'departure_time', '发车时间', '发车'],
        'waybill': ['运单号', 'waybill', '运单号'],
    }
    
    col_map = {}
    row_values = [str(sheet.cell_value(header_row, col)).strip().lower() 
                  for col in range(sheet.ncols)]
    
    for key, keywords in keywords.items():
        for col_idx, val in enumerate(row_values):
            if any(kw.lower() in val for kw in keywords):
                col_map[key] = col_idx
                break
    return col_map


def _map_columns_openpyxl(sheet, header_row: int) -> Dict[str, int]:
    keywords = {
        'container_no': ['箱号', 'container', 'container_no', 'container_no.', '箱号'],
        'container_type': ['箱型', 'container_type', '类型'],
        'chinese_wagon': ['中方车号', '中方车号', 'chinese_wagon'],
        'destination': ['目的站', 'destination', '到站', '到站'],
        'chinese_wagon_no': ['中方车号', '中方车号', 'chinese_wagon'],
        'train_no': ['班列号', 'train_no', 'train_no', '班次'],
        'departure_time': ['发车时间', 'departure_time', '发车时间', '发车'],
        'waybill': ['运单号', 'waybill', '运单号'],
    }
    
    col_map = {}
    for col_idx in range(1, sheet.max_column + 1):
        val = str(sheet.cell(row=header_row, column=col_idx).value or '').strip().lower()
        for key, keywords in keywords.items():
            if any(kw.lower() in val for kw in keywords):
                col_map[key] = col_idx
                break
    return col_map


def _has_required_columns(col_map: Dict[str, int]) -> bool:
    """检查是否包含必要的列"""
    required = ['container_no']  # 至少需要箱号
    return any(k in col_map for k in ['container_no', 'container_no_alt'])


def _parse_row(sheet, row_idx: int, col_map: Dict[str, int]) -> Optional[Dict[str, Any]]:
    """解析 xlrd sheet 的一行数据"""
    def get_cell(key: str) -> str:
        col = col_map.get(key)
        if col is not None:
            val = sheet.cell_value(row_idx, col_map[key])
            return str(val).strip() if val else ""
        return ""
    
    container_no = _get_cell(sheet, col_map, 'container_no', row_idx) or \
                   _get_cell(sheet, col_map, 'container_no_alt', row_idx)
    
    if not container_no:
        return None
        
    return {
        'container_no': container_no,
        'container_type': _get_cell(sheet, col_map, 'container_type', row_idx),
        'chinese_wagon': _get_cell(sheet, col_map, 'chinese_wagon', row_idx),
        'destination': _get_cell(sheet, col_map, 'destination', row_idx),
        'train_no': _get_cell(sheet, col_map, 'train_no', row_idx),
        'departure_time': _get_cell(sheet, col_map, 'departure_time', row_idx),
        'waybill': _get_cell(sheet, col_map, 'waybill', row_idx),
    }


def _get_cell(sheet, col_map: Dict[str, int], key: str, row_idx: int) -> str:
    col = col_map.get(key)
    if col is not None:
        val = sheet.cell_value(row_idx, col)
        return str(val).strip() if val else ""
    return ""


def _get_cell_openpyxl(sheet, col_map: Dict[str, int], key: str, row_idx: int) -> str:
    col = col_map.get(key)
    if col is not None:
        val = sheet.cell(row=row_idx, column=col).value
        return str(val).strip() if val else ""
    return ""


def _parse_row_openpyxl(sheet, row_idx: int, col_map: Dict[str, int]) -> Optional[Dict[str, Any]]:
    container_no = _get_cell_openpyxl(sheet, col_map, 'container_no', row_idx)
    if not container_no:
        return None
        
    return {
        'container_no': container_no,
        'container_type': _get_cell_openpyxl(sheet, col_map, 'container_type', row_idx),
        'chinese_wagon': _get_cell_openpyxl(sheet, col_map, 'chinese_wagon', row_idx),
        'destination': _get_cell_openpyxl(sheet, col_map, 'destination', row_idx),
        'train_no': _get_cell_openpyxl(sheet, col_map, 'train_no', row_idx),
        'departure_time': _get_cell_openpyxl(sheet, col_map, 'departure_time', row_idx),
        'waybill': _get_cell_openpyxl(sheet, col_map, 'waybill', row_idx),
    }


def _parse_row_text(row: List[str], col_map: Dict[str, int]) -> Optional[Dict[str, Any]]:
    """解析文本行"""
    container_no = _get_text_cell(row, col_map, 'container_no')
    if not container_no:
        return None
        
    return {
        'container_no': container_no,
        'container_type': _get_text_cell(row, col_map, 'container_type'),
        'chinese_wagon': _get_text_cell(row, col_map, 'chinese_wagon'),
        'destination': _get_text_cell(col_map, 'destination'),
        'train_no': _get_text_cell(col_map, 'train_no'),
        'departure_time': _get_text_cell(col_map, 'departure_time'),
        'waybill': _get_text_cell(col_map, 'waybill'),
    }


def _get_text_cell(row: List[str], col_map: Dict[str, int], key: str) -> str:
    col = col_map.get(key)
    if col is not None and col < len(row):
        return str(row[col]).strip()
    return ""


# Alias for backward compatibility with Tracing_Robot_IMAP.py
parse_tracing_xls_attachment = parse_tracing_xls