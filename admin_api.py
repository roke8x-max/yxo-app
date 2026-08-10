#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
管理页 API（毛骁洋 全权；杨雅雯/冯茜/韩文豪 受限：仅价格维护）
---------------------------------------------------
功能：
1. 服务状态：本机各服务端口探活 + 邮件机器人最近运行情况 + 最近错误日志
2. 转发日志：ATB / DSK 本地日志、运踪飞书日志表、error_logs 错误日志
3. 配置管理：运踪配置表 / DSK配置表（ATB共用）在线增删改（直连飞书 HTTP API）

说明：
- 飞书凭据从 D:/YXO_DATA/WeComBot/config.py 动态加载（避免与本项目 config.py 重名冲突）
- 配置写入后自动清掉 ATB/DSK 的本地缓存文件，机器人下次运行立即用新配置
- 权限：毛骁洋 全权；杨雅雯/冯茜/韩文豪 为受限管理员，仅价格维护类接口放行（allow_limited=True），服务状态/转发日志/配置管理仍仅毛骁洋。
"""
import os
import json
import glob
import socket
import importlib.util
import urllib.request
import urllib.parse
from datetime import datetime, timedelta
import sqlite3
import re
import config  # 同目录 yxo_app 配置，提供本地库路径 DB_PATH

from flask import Blueprint, request, jsonify, render_template

admin_bp = Blueprint("admin", __name__)

ADMIN_USER = "毛骁洋"

# 受限管理员：可进入系统管理页，但仅能使用「价格维护 / 选项维护 / 回收站」三项。
# 服务状态 / 转发日志 / 配置管理（飞书表）仍仅 ADMIN_USER 可用。
LIMITED_ADMINS = {"杨雅雯", "冯茜", "韩文豪"}

# ---------- 加载外部模块（可配置路径 + 优雅降级，2026-08-04 dev 改造） ----------
# 原实现硬编码 D:\YXO_DATA\{WeComBot,MailBots}\*.py，本机 dev 环境无法 import。
# 新实现：路径可经环境变量 / config_local.py 配置；文件缺失时加载"虚拟模块"降级，
#         函数内所有 wb.XXX / atomic_write_json(...) 调用保持原样即可工作，
#         飞书类接口因凭据为空在本机失败（dev 不调用），生产环境有真文件则完全不变。
try:
    import config_local  # 仅本地存在（.gitignore 屏蔽），生产/CI 无此文件
except ImportError:
    config_local = None


def _resolve_path(env_name, local_attr, default):
    """路径解析优先级：环境变量 > config_local.<attr> > 默认 D 盘路径"""
    if os.environ.get(env_name):
        return os.environ[env_name]
    if config_local is not None and hasattr(config_local, local_attr):
        return getattr(config_local, local_attr)
    return default


def _load_optional_module(name, path):
    """延迟加载外部模块：文件不存在或导入失败时返回 None（由调用方降级处理）。"""
    if not os.path.exists(path):
        print(f"[admin_api] WARNING: 外部模块 {path} 不存在，使用虚拟模块降级（dev 环境预期）")
        return None
    try:
        spec = importlib.util.spec_from_file_location(name, path)
        mod = importlib.util.module_from_spec(spec)
        spec.loader.exec_module(mod)
        return mod
    except Exception as e:
        print(f"[admin_api] WARNING: 外部模块 {path} 加载失败: {e}，使用虚拟模块降级")
        return None


def _make_dummy_wb():
    """虚拟飞书配置模块：所有属性返回空串，避免 wb.XXX 在 dev 环境 AttributeError。"""
    import types
    m = types.SimpleNamespace()
    for _a in ("TABLE_CONFIG", "TABLE_DSK_CONFIG", "TABLE_LOG",
               "FEISHU_APP_ID", "FEISHU_APP_SECRET", "FEISHU_APP_TOKEN"):
        setattr(m, _a, "")
    return m


def _fallback_atomic_write_json(path, data):
    """无 common_io 时的简单实现（dev 环境用，无 Windows 文件占用回退）。"""
    import tempfile, os
    tmp = path + ".tmp"
    with open(tmp, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)
    os.replace(tmp, path)


WECOMBOT_CONFIG_PATH = _resolve_path(
    "WECOMBOT_CONFIG_PATH", "WECOMBOT_CONFIG_PATH", r"D:\YXO_DATA\WeComBot\config.py")
COMMON_IO_PATH = _resolve_path(
    "MAILBOTS_COMMON_IO_PATH", "MAILBOTS_COMMON_IO_PATH", r"D:\YXO_DATA\MailBots\common_io.py")

wb = _load_optional_module("wb_config", WECOMBOT_CONFIG_PATH) or _make_dummy_wb()
_common_io_mod = _load_optional_module("common_io", COMMON_IO_PATH)
atomic_write_json = getattr(_common_io_mod, "atomic_write_json", None) or _fallback_atomic_write_json

FEISHU_BASE = "https://open.feishu.cn/open-apis"

# 可管理的飞书表：key -> (table_id, 可编辑字段列表, 关键字段名)
TABLES = {
    "tracing_config": {
        "table_id": wb.TABLE_CONFIG,
        "fields": ["班列号", "公司名称", "收件人 (To)", "抄送人 (CC)"],
        "label": "运踪配置表",
    },
    "dsk_config": {
        "table_id": wb.TABLE_DSK_CONFIG,
        "fields": ["箱号", "公司名称", "收件人 (To)", "抄送人 (CC)"],
        "label": "DSK/ATB 配置表",
    },
    "tracing_log": {
        "table_id": wb.TABLE_LOG,
        "fields": [],  # 只读
        "label": "运踪日志表",
    },
}

# ATB/DSK 机器人的本地配置缓存（写配置后删除，让机器人立刻拿到新配置）
DSK_CACHE_FILE = r"D:\YXO_DATA\WeComBot\cache\dsk_config_cache.json"

BOT_LOG_DIR = r"D:\YXO_DATA\WeComBot\logs"
ERROR_LOG_DIR = r"D:\YXO_DATA\MailBots\error_logs"
TRACING_REPORT_DIR = r"D:\YXO_DATA\MailBots\TracingLog"

# 机器人统一配置（2026-08-03 姐姐规格书）：config 文件路径映射
BOT_CONFIG_MAP = {
    "draft_forward": r"D:\YXO_DATA\MailBots\draft_robot_config.json",
    "waybill":       r"D:\YXO_DATA\MailBots\waybill_robot_config.json",
    "atb":           r"D:\YXO_DATA\MailBots\atb_robot_config.json",
    "dsk":           r"D:\YXO_DATA\MailBots\dsk_robot_config.json",
    "tracing":       r"D:\YXO_DATA\MailBots\tracing_robot_config.json",
}
# live 缺省值：draft/waybill 默认关闭（保持原语义）；atb/dsk/tracing 默认开启
# （姐姐规格书第四节：这 3 个 config 若文件缺失或 live 缺省 → 视为 true，不误停现有运行）
# 注：数据库同步机器人(syncer)已于 2026-08-03 经毛骁洋确认提前退休（飞书数据已不如本地库权威），不再纳入统一开关。
BOT_LIVE_DEFAULT = {
    "draft_forward": False,
    "waybill": False,
    "atb": True,
    "dsk": True,
    "tracing": True,
}

# 本机常驻服务端口表
SERVICES = [
    ("nginx 反向代理", 5000),
    ("订舱数据管理 (YXO)", 5011),
    ("订舱助手 (企微机器人)", 5001),
    ("FileBrowser 文件", 8080),
    ("Netdata 监控", 19999),
]
# 注：AList/Navidrome/Jellyfin 为个人媒体服务，已从管理页监控移除（数科部要求公网不暴露）。
# 其存活状态改由导航页 /index 以"状态卡片"展示（轮询 /api/service_status），链接不对外。


# ====================== 权限 ======================
def _check_user(allow_limited=False):
    u = request.args.get("user") or (request.get_json(silent=True) or {}).get("user", "")
    if allow_limited:
        return u == ADMIN_USER or u in LIMITED_ADMINS
    return u == ADMIN_USER


def _forbid():
    return jsonify({"ok": False, "msg": "无权限：该功能仅限管理员使用"}), 403


# ====================== 飞书 HTTP 封装 ======================
_token_cache = {"token": "", "expire_at": 0}


def _feishu_token():
    now = datetime.now().timestamp()
    if _token_cache["token"] and now < _token_cache["expire_at"] - 60:
        return _token_cache["token"]
    body = json.dumps({"app_id": wb.FEISHU_APP_ID, "app_secret": wb.FEISHU_APP_SECRET}).encode("utf-8")
    req = urllib.request.Request(
        FEISHU_BASE + "/auth/v3/tenant_access_token/internal",
        data=body, method="POST", headers={"Content-Type": "application/json"})
    with urllib.request.urlopen(req, timeout=15) as resp:
        data = json.load(resp)
    if data.get("code") != 0:
        raise RuntimeError("获取飞书凭证失败: " + str(data))
    _token_cache["token"] = data["tenant_access_token"]
    _token_cache["expire_at"] = now + int(data.get("expire", 3600))
    return _token_cache["token"]


def _feishu_req(method, path, payload=None, params=None):
    url = FEISHU_BASE + path
    if params:
        url += "?" + urllib.parse.urlencode(params)
    body = json.dumps(payload, ensure_ascii=False).encode("utf-8") if payload is not None else None
    req = urllib.request.Request(url, data=body, method=method, headers={
        "Content-Type": "application/json",
        "Authorization": "Bearer " + _feishu_token(),
    })
    with urllib.request.urlopen(req, timeout=30) as resp:
        return json.load(resp)


def _feishu_list_records(table_id):
    """拉取全表记录（带创建时间）"""
    items, page_token = [], ""
    while True:
        params = {"page_size": 500, "automatic_fields": "true"}
        if page_token:
            params["page_token"] = page_token
        data = _feishu_req("GET", f"/bitable/v1/apps/{wb.FEISHU_APP_TOKEN}/tables/{table_id}/records", params=params)
        if data.get("code") != 0:
            raise RuntimeError("读取飞书表失败: " + str(data.get("msg")))
        d = data.get("data") or {}
        items.extend(d.get("items") or [])
        if not d.get("has_more"):
            break
        page_token = d.get("page_token", "")
    return items


def _clear_dsk_cache():
    try:
        if os.path.exists(DSK_CACHE_FILE):
            os.remove(DSK_CACHE_FILE)
    except Exception:
        pass


# ====================== 页面 ======================
@admin_bp.route("/admin")
def admin_page():
    # 禁止浏览器缓存：避免受限管理员改版后一直看到旧的 admin.html
    from flask import make_response
    resp = make_response(render_template("admin.html"))
    resp.headers["Cache-Control"] = "no-store, no-cache, must-revalidate, max-age=0"
    resp.headers["Pragma"] = "no-cache"
    resp.headers["Expires"] = "0"
    return resp


# ====================== 服务状态 ======================
def _port_alive(port, timeout=1.5):
    try:
        with socket.create_connection(("127.0.0.1", port), timeout=timeout):
            return True
    except Exception:
        return False


def _latest_file_info(pattern):
    """返回匹配文件中最新一个的 (文件名, 修改时间str)；没有则 (None, None)"""
    files = glob.glob(pattern)
    if not files:
        return None, None
    f = max(files, key=os.path.getmtime)
    ts = datetime.fromtimestamp(os.path.getmtime(f)).strftime("%Y-%m-%d %H:%M:%S")
    return os.path.basename(f), ts


@admin_bp.route("/api/admin/status")
def api_status():
    if not _check_user():
        return _forbid()
    services = [{"name": n, "port": p, "alive": _port_alive(p)} for n, p in SERVICES]

    # 邮件机器人（非常驻，按最近日志时间判断活跃度）
    bots = []
    for key, label, pattern in [
        ("draft_forward", "草单转发机器人", os.path.join(BOT_LOG_DIR, "draft_forward_*.log")),
        ("waybill",       "运单号机器人",   os.path.join(r"D:\YXO_DATA\MailBots\logs", "waybill_*.log")),
        ("atb",           "ATB 转发机器人",  os.path.join(BOT_LOG_DIR, "atb_*.log")),
        ("dsk",           "DSK 转发机器人",  os.path.join(BOT_LOG_DIR, "dsk_*.log")),
        ("tracing",       "运踪转发机器人",  os.path.join(TRACING_REPORT_DIR, "*.txt")),
    ]:
        fname, ts = _latest_file_info(pattern)
        # 可选增强：读 config 的 live 显示「运行中 / 已停用」
        live = None
        cfg_path = BOT_CONFIG_MAP.get(key)
        if cfg_path:
            try:
                with open(cfg_path, "r", encoding="utf-8") as f:
                    c = json.load(f)
                live = bool(c.get("live", BOT_LIVE_DEFAULT.get(key, False)))
            except Exception:
                live = BOT_LIVE_DEFAULT.get(key, False)
        bots.append({"key": key, "name": label, "last_file": fname,
                     "last_time": ts, "live": live})

    # 最近 48 小时错误日志
    errors = []
    cutoff = datetime.now() - timedelta(hours=48)
    for f in sorted(glob.glob(os.path.join(ERROR_LOG_DIR, "*.log")), key=os.path.getmtime, reverse=True):
        mt = datetime.fromtimestamp(os.path.getmtime(f))
        if mt < cutoff:
            break
        errors.append({"file": os.path.basename(f), "time": mt.strftime("%Y-%m-%d %H:%M:%S")})

    return jsonify({"ok": True, "services": services, "bots": bots,
                    "recent_errors": errors,
                    "checked_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S")})


# ====================== 日志查看 ======================
def _tail_file(path, lines=300):
    try:
        with open(path, "r", encoding="utf-8", errors="replace") as f:
            all_lines = f.readlines()
        return "".join(all_lines[-lines:])
    except FileNotFoundError:
        return ""


@admin_bp.route("/api/admin/logs")
def api_logs():
    if not _check_user():
        return _forbid()
    bot = request.args.get("bot", "atb")
    date = request.args.get("date") or datetime.now().strftime("%Y-%m-%d")
    lines = min(int(request.args.get("lines", 300)), 2000)

    if bot in ("atb", "dsk"):
        path = os.path.join(BOT_LOG_DIR, f"{bot}_{date}.log")
        content = _tail_file(path, lines)
        # 列出可选日期
        dates = sorted({os.path.basename(f)[len(bot) + 1:-4]
                        for f in glob.glob(os.path.join(BOT_LOG_DIR, f"{bot}_*.log"))}, reverse=True)
        return jsonify({"ok": True, "type": "text", "content": content or "（该日期暂无日志）", "dates": dates})

    if bot == "tracing":
        # 运踪转发日志：本地 yxo.db tracing_log 表（P3 起为唯一数据源，替代飞书 TABLE_LOG）
        # Task#174：分页 + 公司/班列 搜索，避免一次性渲染全部行卡顿
        try:
            company = (request.args.get("company") or "").strip()
            train = (request.args.get("train") or "").strip()
            page = max(int(request.args.get("page", 1)), 1)
            page_size = min(max(int(request.args.get("page_size", 50)), 1), 200)

            where, params = [], []
            if company:
                where.append("company LIKE ?")
                params.append(f"%{company}%")
            if train:
                where.append("train_no LIKE ?")
                params.append(f"%{train}%")
            wh = (" WHERE " + " AND ".join(where)) if where else ""

            conn = sqlite3.connect(config.DB_PATH)
            conn.row_factory = sqlite3.Row
            total = conn.execute(f"SELECT COUNT(*) FROM tracing_log{wh}", params).fetchone()[0]
            offset = (page - 1) * page_size
            cur = conn.execute(
                'SELECT log_date AS "日期", train_no AS "班列号", company AS "接收公司", forward_detail AS "转发详情" '
                f'FROM tracing_log{wh} ORDER BY log_date DESC LIMIT ? OFFSET ?',
                params + [page_size, offset]
            )
            rows = [dict(r) for r in cur.fetchall()]
            # 下拉候选：去重公司 / 班列
            comps = [r[0] for r in conn.execute(
                "SELECT DISTINCT company FROM tracing_log WHERE company<>'' ORDER BY company").fetchall()]
            trains = [r[0] for r in conn.execute(
                "SELECT DISTINCT train_no FROM tracing_log WHERE train_no<>'' ORDER BY train_no DESC").fetchall()]
            conn.close()
            pages = (total + page_size - 1) // page_size if total else 1
            return jsonify({"ok": True, "type": "table",
                            "columns": ["日期", "班列号", "接收公司", "转发详情"],
                            "rows": rows, "total": total, "page": page,
                            "page_size": page_size, "pages": pages,
                            "companies": comps, "trains": trains})
        except Exception as e:
            return jsonify({"ok": False, "msg": "读取本地运踪日志失败：" + str(e)})

    if bot == "errors":
        file = request.args.get("file", "")
        if file:
            # 防目录穿越：只允许纯文件名
            if os.path.basename(file) != file:
                return jsonify({"ok": False, "msg": "非法文件名"})
            content = _tail_file(os.path.join(ERROR_LOG_DIR, file), lines)
            return jsonify({"ok": True, "type": "text", "content": content or "（文件为空）"})
        files = sorted(glob.glob(os.path.join(ERROR_LOG_DIR, "*.log")), key=os.path.getmtime, reverse=True)
        return jsonify({"ok": True, "type": "list",
                        "files": [{"file": os.path.basename(f),
                                   "time": datetime.fromtimestamp(os.path.getmtime(f)).strftime("%Y-%m-%d %H:%M")}
                                  for f in files[:100]]})

    return jsonify({"ok": False, "msg": "未知日志类型"})


# ====================== 配置管理（飞书表增删改查） ======================
def _get_table_or_none(key, need_editable=False):
    t = TABLES.get(key)
    if not t:
        return None
    if need_editable and not t["fields"]:
        return None
    return t


# ====================== 本地配置表闭环（yxo.db bot_config，替代飞书） ======================
# 背景：P4 已把路由数据从飞书迁到 yxo.db，机器人改读 yxo.db；但配置管理页仍写飞书导致闭环断裂。
# 这里把 tracing_config / dsk_config 的增删改查改走 yxo.db，UI 加的路由机器人下次运行即生效。
# 模型：tracing -> bot='tracking'（scope='train' 存 班列号->[公司]，scope='company' 存 公司->收件人）；
#       dsk    -> bot in ('dsk','atb') scope='company'（DSK/ATB 共用历史表，写两边）。
DB_TABLES = {
    "tracing_config": {"label": "运踪配置表（按班列号）", "kind": "tracing",
                       "fields": ["班列号", "公司名称", "收件人 (To)", "抄送人 (CC)"]},
    "dsk_config":     {"label": "DSK/ATB 配置表（按公司）", "kind": "dsk",
                       "fields": ["公司名称", "收件人 (To)", "抄送人 (CC)"]},
}


def _db_join_addrs(s):
    try:
        return ";".join(json.loads(s or "[]"))
    except Exception:
        return (s or "")


def _db_parse_addrs(s):
    if s is None:
        return []
    s = str(s).strip()
    if not s:
        return []
    return [p.strip() for p in re.split(r"[;,]", s) if p.strip()]


def _db_upsert_company(cur, bot, company, to_list, cc_list):
    cur.execute(
        """INSERT INTO bot_config(bot,scope,key,to_addrs,cc_addrs,source,updated_at)
           VALUES(?,?,?,?,?,'manual',datetime('now','localtime'))
           ON CONFLICT(bot,scope,key) DO UPDATE SET
             to_addrs=excluded.to_addrs, cc_addrs=excluded.cc_addrs,
             source='manual', updated_at=excluded.updated_at""",
        (bot, "company", company,
         json.dumps(to_list, ensure_ascii=False), json.dumps(cc_list, ensure_ascii=False)))


def _db_cfg_list(kind):
    conn = sqlite3.connect(config.DB_PATH); cur = conn.cursor()
    rows = []
    if kind == "tracing":
        comp = {k: (to, cc) for k, to, cc in cur.execute(
            "SELECT key,to_addrs,cc_addrs FROM bot_config WHERE bot='tracking' AND scope='company'")}
        for tid, ex in cur.execute(
                "SELECT key,extra FROM bot_config WHERE bot='tracking' AND scope='train' ORDER BY CAST(key AS INT)"):
            try:
                companies = json.loads(ex or "{}").get("companies", [])
            except Exception:
                companies = []
            for c in companies:
                to, cc = comp.get(c, ("[]", "[]"))
                rows.append({"record_id": "TR|%s|%s" % (tid, c), "班列号": tid, "公司名称": c,
                             "收件人 (To)": _db_join_addrs(to), "抄送人 (CC)": _db_join_addrs(cc)})
    else:
        for k, to, cc in cur.execute(
                "SELECT key,to_addrs,cc_addrs FROM bot_config WHERE bot='dsk' AND scope='company' ORDER BY key"):
            rows.append({"record_id": "DSK|%s" % k, "公司名称": k,
                         "收件人 (To)": _db_join_addrs(to), "抄送人 (CC)": _db_join_addrs(cc)})
    conn.close()
    return rows


def _db_cfg_upsert_tracing(train, company, to_list, cc_list):
    conn = sqlite3.connect(config.DB_PATH); cur = conn.cursor()
    _db_upsert_company(cur, "tracking", company, to_list, cc_list)
    row = cur.execute("SELECT extra FROM bot_config WHERE bot='tracking' AND scope='train' AND key=?",
                      (train,)).fetchone()
    if row:
        try:
            comps = json.loads(row[0] or "{}").get("companies", [])
        except Exception:
            comps = []
        if company not in comps:
            comps.append(company)
        cur.execute("UPDATE bot_config SET extra=?, updated_at=datetime('now','localtime') "
                    "WHERE bot='tracking' AND scope='train' AND key=?",
                    (json.dumps({"companies": comps}, ensure_ascii=False), train))
    else:
        cur.execute("INSERT INTO bot_config(bot,scope,key,to_addrs,cc_addrs,extra,source,updated_at) "
                    "VALUES('tracking','train',?,'[]','[]',?,'manual',datetime('now','localtime'))",
                    (train, json.dumps({"companies": [company]}, ensure_ascii=False)))
    conn.commit(); conn.close()


def _db_cfg_upsert_dsk(company, to_list, cc_list):
    conn = sqlite3.connect(config.DB_PATH); cur = conn.cursor()
    _db_upsert_company(cur, "dsk", company, to_list, cc_list)
    _db_upsert_company(cur, "atb", company, to_list, cc_list)   # DSK/ATB 共用，写两边
    conn.commit(); conn.close()
    _clear_dsk_cache()


def _db_cfg_delete(kind, record_id):
    conn = sqlite3.connect(config.DB_PATH); cur = conn.cursor()
    if kind == "tracing":
        parts = record_id.split("|")
        if len(parts) == 3:
            _, train, company = parts
            row = cur.execute("SELECT extra FROM bot_config WHERE bot='tracking' AND scope='train' AND key=?",
                              (train,)).fetchone()
            if row:
                try:
                    comps = json.loads(row[0] or "{}").get("companies", [])
                except Exception:
                    comps = []
                comps = [c for c in comps if c != company]
                if comps:
                    cur.execute("UPDATE bot_config SET extra=?, updated_at=datetime('now','localtime') "
                                "WHERE bot='tracking' AND scope='train' AND key=?",
                                (json.dumps({"companies": comps}, ensure_ascii=False), train))
                else:
                    cur.execute("DELETE FROM bot_config WHERE bot='tracking' AND scope='train' AND key=?",
                                (train,))
    else:
        parts = record_id.split("|")
        if len(parts) == 2:
            _, company = parts
            cur.execute("DELETE FROM bot_config WHERE bot IN ('dsk','atb') AND scope='company' AND key=?",
                        (company,))
            _clear_dsk_cache()
    conn.commit(); conn.close()


@admin_bp.route("/api/admin/table/<key>", methods=["GET"])
def api_table_list(key):
    if not _check_user():
        return _forbid()
    t = _get_table_or_none(key)
    if not t:
        return jsonify({"ok": False, "msg": "未知配置表"}), 404
    if key in DB_TABLES:                      # 本地 yxo.db 闭环（替代飞书）
        try:
            rows = _db_cfg_list(DB_TABLES[key]["kind"])
        except Exception as e:
            return jsonify({"ok": False, "msg": "读取失败：" + str(e)})
        return jsonify({"ok": True, "label": DB_TABLES[key]["label"],
                        "fields": DB_TABLES[key]["fields"], "rows": rows})
    try:
        items = _feishu_list_records(t["table_id"])
    except Exception as e:
        return jsonify({"ok": False, "msg": "读取失败：" + str(e)})
    rows = []
    for it in items:
        f = it.get("fields") or {}
        row = {"record_id": it.get("record_id")}
        for name in (t["fields"] or list(f.keys())):
            v = f.get(name)
            # 飞书文本字段可能是富文本数组，统一转成字符串
            if isinstance(v, list):
                v = "".join(x.get("text", "") if isinstance(x, dict) else str(x) for x in v)
            row[name] = "" if v is None else str(v)
        ct = it.get("created_time")
        if ct:
            try:
                row["_created"] = datetime.fromtimestamp(float(ct) / 1000).strftime("%Y-%m-%d")
            except Exception:
                row["_created"] = ""
        rows.append(row)
    return jsonify({"ok": True, "label": t["label"], "fields": t["fields"], "rows": rows})


@admin_bp.route("/api/admin/table/<key>", methods=["POST"])
def api_table_create(key):
    if not _check_user():
        return _forbid()
    t = _get_table_or_none(key, need_editable=True)
    if not t:
        return jsonify({"ok": False, "msg": "该表不可编辑"}), 400
    data = request.get_json(force=True, silent=True) or {}
    if key in DB_TABLES:                      # 本地 yxo.db 闭环（替代飞书）
        try:
            f = {k: str(v).strip() for k, v in (data.get("fields") or {}).items()}
            kind = DB_TABLES[key]["kind"]
            if kind == "tracing":
                train = f.get("班列号", ""); company = f.get("公司名称", "")
                if not train or not company:
                    return jsonify({"ok": False, "msg": "班列号与公司名称必填"}), 400
                _db_cfg_upsert_tracing(train, company,
                                      _db_parse_addrs(f.get("收件人 (To)", "")),
                                      _db_parse_addrs(f.get("抄送人 (CC)", "")))
            else:
                company = f.get("公司名称", "")
                if not company:
                    return jsonify({"ok": False, "msg": "公司名称必填"}), 400
                _db_cfg_upsert_dsk(company,
                                   _db_parse_addrs(f.get("收件人 (To)", "")),
                                   _db_parse_addrs(f.get("抄送人 (CC)", "")))
            return jsonify({"ok": True})
        except Exception as e:
            return jsonify({"ok": False, "msg": "写入异常：" + str(e)})
    fields = {k: str(v) for k, v in (data.get("fields") or {}).items() if k in t["fields"]}
    if not fields:
        return jsonify({"ok": False, "msg": "没有可写入的字段"}), 400
    try:
        r = _feishu_req("POST", f"/bitable/v1/apps/{wb.FEISHU_APP_TOKEN}/tables/{t['table_id']}/records",
                        payload={"fields": fields})
        if r.get("code") != 0:
            return jsonify({"ok": False, "msg": "飞书写入失败：" + str(r.get("msg"))})
    except Exception as e:
        return jsonify({"ok": False, "msg": "写入异常：" + str(e)})
    if key == "dsk_config":
        _clear_dsk_cache()
    return jsonify({"ok": True})


@admin_bp.route("/api/admin/table/<key>/<record_id>", methods=["PATCH"])
def api_table_update(key, record_id):
    if not _check_user():
        return _forbid()
    t = _get_table_or_none(key, need_editable=True)
    if not t:
        return jsonify({"ok": False, "msg": "该表不可编辑"}), 400
    data = request.get_json(force=True, silent=True) or {}
    if key in DB_TABLES:                      # 本地 yxo.db 闭环（替代飞书）
        try:
            f = {k: str(v).strip() for k, v in (data.get("fields") or {}).items()}
            kind = DB_TABLES[key]["kind"]
            to = _db_parse_addrs(f.get("收件人 (To)", ""))
            cc = _db_parse_addrs(f.get("抄送人 (CC)", ""))
            if kind == "tracing":
                parts = record_id.split("|")
                if len(parts) != 3:
                    return jsonify({"ok": False, "msg": "非法记录标识"}), 400
                _, train, old_company = parts
                new_company = (f.get("公司名称", "") or old_company).strip()
                conn = sqlite3.connect(config.DB_PATH); cur = conn.cursor()
                _db_upsert_company(cur, "tracking", new_company, to, cc)
                if new_company != old_company:
                    row = cur.execute("SELECT extra FROM bot_config WHERE bot='tracking' AND scope='train' AND key=?", (train,)).fetchone()
                    comps = json.loads(row[0] or "{}").get("companies", []) if row else []
                    comps = [c for c in comps if c != old_company]
                    if new_company not in comps:
                        comps.append(new_company)
                    if comps:
                        cur.execute("UPDATE bot_config SET extra=?, updated_at=datetime('now','localtime') WHERE bot='tracking' AND scope='train' AND key=?",
                                    (json.dumps({"companies": comps}, ensure_ascii=False), train))
                    else:
                        cur.execute("DELETE FROM bot_config WHERE bot='tracking' AND scope='train' AND key=?", (train,))
                conn.commit(); conn.close()
            else:
                company = (f.get("公司名称", "") or record_id.split("|")[-1]).strip()
                _db_cfg_upsert_dsk(company, to, cc)
            return jsonify({"ok": True})
        except Exception as e:
            return jsonify({"ok": False, "msg": "更新异常：" + str(e)})
    fields = {k: str(v) for k, v in (data.get("fields") or {}).items() if k in t["fields"]}
    if not fields:
        return jsonify({"ok": False, "msg": "没有可写入的字段"}), 400
    try:
        r = _feishu_req("PUT", f"/bitable/v1/apps/{wb.FEISHU_APP_TOKEN}/tables/{t['table_id']}/records/{record_id}",
                        payload={"fields": fields})
        if r.get("code") != 0:
            return jsonify({"ok": False, "msg": "飞书更新失败：" + str(r.get("msg"))})
    except Exception as e:
        return jsonify({"ok": False, "msg": "更新异常：" + str(e)})
    if key == "dsk_config":
        _clear_dsk_cache()
    return jsonify({"ok": True})


@admin_bp.route("/api/admin/table/<key>/<record_id>", methods=["DELETE"])
def api_table_delete(key, record_id):
    if not _check_user():
        return _forbid()
    t = _get_table_or_none(key, need_editable=True)
    if not t:
        return jsonify({"ok": False, "msg": "该表不可编辑"}), 400
    if key in DB_TABLES:                      # 本地 yxo.db 闭环（替代飞书）
        try:
            _db_cfg_delete(DB_TABLES[key]["kind"], record_id)
            return jsonify({"ok": True})
        except Exception as e:
            return jsonify({"ok": False, "msg": "删除异常：" + str(e)})
    try:
        r = _feishu_req("DELETE", f"/bitable/v1/apps/{wb.FEISHU_APP_TOKEN}/tables/{t['table_id']}/records/{record_id}")
        if r.get("code") != 0:
            return jsonify({"ok": False, "msg": "飞书删除失败：" + str(r.get("msg"))})
    except Exception as e:
        return jsonify({"ok": False, "msg": "删除异常：" + str(e)})
    if key == "dsk_config":
        _clear_dsk_cache()
    return jsonify({"ok": True})


# ====================== 草单转发机器人配置（draft_robot_config.json） ======================
DRAFT_ROBOT_CFG = r"D:\YXO_DATA\MailBots\draft_robot_config.json"


def _draft_robot_last_run():
    try:
        fs = sorted(glob.glob(os.path.join(BOT_LOG_DIR, "draft_forward_*.log")),
                    key=os.path.getmtime, reverse=True)
        if fs:
            return datetime.fromtimestamp(os.path.getmtime(fs[0])).strftime("%Y-%m-%d %H:%M:%S")
    except Exception:
        pass
    return None


@admin_bp.route("/api/admin/draft_robot_config", methods=["GET"])
def api_draft_robot_config():
    if not _check_user():  # 配置管理仅毛骁洋
        return _forbid()
    default_accounts = [
        "maoxiaoyang@cqtransit.com", "yangyawen@cqtransit.com",
        "fengqian@cqtransit.com", "hanwenhao@cqtransit.com",
    ]
    try:
        with open(DRAFT_ROBOT_CFG, "r", encoding="utf-8") as f:
            cfg = json.load(f)
    except Exception:
        cfg = {}
    return jsonify({"ok": True,
                    "live": bool(cfg.get("live", False)),
                    "forward_since": cfg.get("forward_since"),
                    "accounts": cfg.get("accounts", default_accounts),
                    "last_run": _draft_robot_last_run()})


@admin_bp.route("/api/admin/draft_robot_config", methods=["POST"])
def api_draft_robot_config_save():
    if not _check_user():
        return _forbid()
    data = request.get_json(silent=True) or {}
    if (data.get("user") or "") != ADMIN_USER:
        return _forbid()
    try:
        live = bool(data.get("live", False))
        fs = data.get("forward_since")  # "YYYY-MM-DDTHH:MM:SS" 或 None
        if fs:
            datetime.strptime(fs, "%Y-%m-%dT%H:%M:%S")  # 校验
        accounts = [a.strip().lower() for a in (data.get("accounts") or []) if a.strip()]
        cfg = {"live": live, "forward_since": fs, "accounts": accounts}
        os.makedirs(os.path.dirname(DRAFT_ROBOT_CFG), exist_ok=True)
        atomic_write_json(DRAFT_ROBOT_CFG, cfg)
    except ValueError:
        return jsonify({"ok": False, "msg": "转发起始时间格式不正确（应为 YYYY-MM-DDTHH:MM:SS）"}), 400
    except Exception as e:
        return jsonify({"ok": False, "msg": "保存失败：" + str(e)}), 500
    return jsonify({"ok": True})


# ====================== 统一机器人配置（所有机器人共用 schema） ======================
# schema：{"live": bool, "forward_since": str|null, "accounts": []}
# 草单转发旧接口 /api/admin/draft_robot_config 保留兼容，新 UI 统一走这里。
@admin_bp.route("/api/admin/bot_config/<name>", methods=["GET"])
def api_bot_config(name):
    if not _check_user():                       # 配置管理仅毛骁洋
        return _forbid()
    path = BOT_CONFIG_MAP.get(name)
    if not path:
        return jsonify(ok=False, msg="未知机器人"), 404
    cfg = {}
    try:
        with open(path, "r", encoding="utf-8") as f:
            cfg = json.load(f)
    except Exception:
        pass
    live_default = BOT_LIVE_DEFAULT.get(name, False)
    return jsonify(ok=True, name=name,
                   live=bool(cfg.get("live", live_default)),
                   forward_since=cfg.get("forward_since"),
                   accounts=cfg.get("accounts", []))


@admin_bp.route("/api/admin/bot_config/<name>", methods=["POST"])
def api_bot_config_save(name):
    if not _check_user():
        return _forbid()
    data = request.get_json(silent=True) or {}
    if (data.get("user") or "") != ADMIN_USER:
        return _forbid()
    path = BOT_CONFIG_MAP.get(name)
    if not path:
        return jsonify(ok=False, msg="未知机器人"), 404
    try:
        live = bool(data.get("live", False))
        fs = data.get("forward_since")
        if fs:
            datetime.strptime(fs, "%Y-%m-%dT%H:%M:%S")   # 校验
        accounts = [a.strip().lower() for a in (data.get("accounts") or []) if a.strip()]
        cfg = {"live": live, "forward_since": fs, "accounts": accounts}
        os.makedirs(os.path.dirname(path), exist_ok=True)
        # 原子写（atomic_write_json 内置 Windows 文件占用回退，等价于原 open(path,'w')+回退方案）
        atomic_write_json(path, cfg)
    except ValueError:
        return jsonify(ok=False, msg="转发起始时间格式不正确（应为 YYYY-MM-DDTHH:MM:SS）"), 400
    except Exception as e:
        return jsonify(ok=False, msg="保存失败：" + str(e)), 500
    return jsonify(ok=True)


# ====================== 价格维护（price_config.json） ======================
PRICE_CONFIG = r"D:\YXO_DATA\MailBots\price_config.json"
PRICE_BACKUP_DIR = r"D:\YXO_DATA\MailBots\price_backups"
PRICE_PENDING = r"D:\YXO_DATA\MailBots\price_import_pending.json"
YXO_DB = r"D:\YXO_DATA\yxo_app\data\yxo.db"

# 与 app.py compute_price 一致的回退站（此处独立实现，避免循环导入）
_MOSCOW_FB = {"别雷拉斯特", "沃尔西诺", "电煤", "谢利亚季诺"}
_MINSK_FB = {"科里亚季奇"}

_MONTH_RE_STR = r"^\d{4}-\d{2}$"


def _load_price_cfg():
    try:
        # utf-8-sig：兼容记事本另存为带来的 BOM（体检报告 3.6）
        with open(PRICE_CONFIG, "r", encoding="utf-8-sig") as f:
            return json.load(f)
    except Exception:
        return {"prices": {}, "exrates": {}}


def _save_price_cfg(cfg):
    """写前自动备份到 price_backups/，再原子替换。"""
    import shutil
    os.makedirs(PRICE_BACKUP_DIR, exist_ok=True)
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    if os.path.exists(PRICE_CONFIG):
        shutil.copy2(PRICE_CONFIG, os.path.join(PRICE_BACKUP_DIR, f"price_config_{ts}.json"))
    atomic_write_json(PRICE_CONFIG, cfg)


def _num_or_none(v):
    if v is None or str(v).strip() == "":
        return None
    try:
        return float(str(v).replace(",", "").replace("$", "").replace("￥", "").strip())
    except Exception:
        return None


def _compute_price_local(port, dest, cntr_type, mk, cfg):
    """按月 key 直接算价（与 app.py compute_price 同逻辑，mk 已解析好）。"""
    if not (port and dest and mk):
        return None
    ct = "COC" if cntr_type and "COC" in str(cntr_type).upper() else "SOC"
    prices = cfg.get("prices", {}).get(mk, {})
    if not prices:
        return None
    pi = prices.get(f"{port}|{dest}")
    if pi is None and dest in _MOSCOW_FB:
        pi = prices.get(f"{port}|莫斯科")
    if pi is None and dest in _MINSK_FB:
        pi = prices.get(f"{port}|明斯克")
    if pi is None:
        return None
    rp = pi.get(ct)
    if rp is None and ct == "COC":
        rp = pi.get("SOC")
    if rp is None:
        return None
    cur = pi.get("currency", "USD")
    if cur == "RMB":
        return round(rp, 2)
    rate = cfg.get("exrates", {}).get(mk, {}).get("USD")
    if rate is None:
        return None
    return round(rp * rate, 2)


def _mk_of_depart(depart_str):
    if not depart_str:
        return None
    head = str(depart_str).strip().split(" ")[0].split("T")[0]
    for fmt in ("%Y/%m/%d", "%Y-%m-%d"):
        try:
            return datetime.strptime(head, fmt).strftime("%Y-%m")
        except Exception:
            continue
    return None


@admin_bp.route("/api/admin/price")
def api_price_get():
    if not _check_user(allow_limited=True):
        return _forbid()
    import re as _re
    cfg = _load_price_cfg()
    months = sorted(set(list(cfg.get("prices", {}).keys()) + list(cfg.get("exrates", {}).keys())), reverse=True)
    month = request.args.get("month") or (months[0] if months else datetime.now().strftime("%Y-%m"))
    if not _re.match(_MONTH_RE_STR, month):
        return jsonify({"ok": False, "msg": "月份格式应为 YYYY-MM"})
    entries = []
    for key, pi in sorted((cfg.get("prices", {}).get(month) or {}).items()):
        port, _, dest = key.partition("|")
        entries.append({"port": port, "dest": dest, "SOC": pi.get("SOC"),
                        "COC": pi.get("COC"), "currency": pi.get("currency", "USD")})
    return jsonify({"ok": True, "months": months, "month": month,
                    "exrates": cfg.get("exrates", {}).get(month, {}),
                    "entries": entries,
                    "unmatched": cfg.get("unmatched", [])})


@admin_bp.route("/api/admin/price/exrate", methods=["POST"])
def api_price_exrate():
    if not _check_user(allow_limited=True):
        return _forbid()
    import re as _re
    d = request.get_json(force=True, silent=True) or {}
    month, cur = str(d.get("month", "")).strip(), str(d.get("currency", "USD")).strip().upper() or "USD"
    rate = _num_or_none(d.get("rate"))
    if not _re.match(_MONTH_RE_STR, month):
        return jsonify({"ok": False, "msg": "月份格式应为 YYYY-MM"})
    if rate is None or rate <= 0:
        return jsonify({"ok": False, "msg": "汇率必须是正数"})
    cfg = _load_price_cfg()
    cfg.setdefault("exrates", {}).setdefault(month, {})[cur] = rate
    _save_price_cfg(cfg)
    return jsonify({"ok": True})


@admin_bp.route("/api/admin/price/entry", methods=["POST"])
def api_price_entry_save():
    """新增或修改一条价格（month+port+dest 定位）。"""
    if not _check_user(allow_limited=True):
        return _forbid()
    import re as _re
    d = request.get_json(force=True, silent=True) or {}
    month = str(d.get("month", "")).strip()
    port = str(d.get("port", "")).strip()
    dest = str(d.get("dest", "")).strip()
    if not _re.match(_MONTH_RE_STR, month) or not port or not dest:
        return jsonify({"ok": False, "msg": "月份/口岸/目的站必填，月份格式 YYYY-MM"})
    soc, coc = _num_or_none(d.get("SOC")), _num_or_none(d.get("COC"))
    if soc is None and coc is None:
        return jsonify({"ok": False, "msg": "SOC/COC 至少填一个价"})
    cur = str(d.get("currency", "USD")).strip().upper() or "USD"
    if cur not in ("USD", "RMB"):
        return jsonify({"ok": False, "msg": "币种只能是 USD 或 RMB"})
    cfg = _load_price_cfg()
    pi = {}
    if soc is not None:
        pi["SOC"] = soc
    if coc is not None:
        pi["COC"] = coc
    pi["currency"] = cur
    cfg.setdefault("prices", {}).setdefault(month, {})[f"{port}|{dest}"] = pi
    _save_price_cfg(cfg)
    return jsonify({"ok": True})


@admin_bp.route("/api/admin/price/entry/delete", methods=["POST"])
def api_price_entry_delete():
    if not _check_user(allow_limited=True):
        return _forbid()
    d = request.get_json(force=True, silent=True) or {}
    month = str(d.get("month", "")).strip()
    key = f"{str(d.get('port','')).strip()}|{str(d.get('dest','')).strip()}"
    cfg = _load_price_cfg()
    if key in (cfg.get("prices", {}).get(month) or {}):
        del cfg["prices"][month][key]
        if not cfg["prices"][month]:
            del cfg["prices"][month]
        _save_price_cfg(cfg)
        return jsonify({"ok": True})
    return jsonify({"ok": False, "msg": "未找到该条价格"})


def _parse_price_workbook(stream):
    """解析上传的《价格表导入模板》。返回 (prices, exrates, pending, errors)
    prices: {month: {"口岸|目的站": {SOC, COC, currency}}}
    pending: 备注含「待确认」的行（不入正式价格）
    errors: 格式非法的行说明"""
    import io
    import re as _re
    import openpyxl
    wbk = openpyxl.load_workbook(io.BytesIO(stream), read_only=True, data_only=True)
    prices, exrates, pending, errors = {}, {}, [], []

    if "价格" in wbk.sheetnames:
        ws = wbk["价格"]
        for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            if row is None or not any(v not in (None, "") for v in row):
                continue
            vals = list(row) + [None] * 7
            mk = str(vals[0] or "").strip()
            port = str(vals[1] or "").strip()
            dest = str(vals[2] or "").strip()
            soc, coc = _num_or_none(vals[3]), _num_or_none(vals[4])
            cur = (str(vals[5] or "").strip().upper() or "USD")
            note = str(vals[6] or "").strip()
            if not (mk or port or dest):
                continue
            if "待确认" in note:
                pending.append({"row": i, "month": mk, "port": port, "dest": dest,
                                "SOC": soc, "COC": coc, "currency": cur, "note": note})
                continue
            if not _re.match(_MONTH_RE_STR, mk):
                errors.append(f"价格 第{i}行：月份「{mk}」格式应为 YYYY-MM")
                continue
            if not port or not dest:
                errors.append(f"价格 第{i}行：口岸/目的站不能为空")
                continue
            if soc is None and coc is None:
                errors.append(f"价格 第{i}行：SOC/COC 至少要有一个价")
                continue
            if cur not in ("USD", "RMB"):
                errors.append(f"价格 第{i}行：币种「{cur}」只能是 USD 或 RMB")
                continue
            pi = {}
            if soc is not None:
                pi["SOC"] = soc
            if coc is not None:
                pi["COC"] = coc
            pi["currency"] = cur
            prices.setdefault(mk, {})[f"{port}|{dest}"] = pi

    if "汇率" in wbk.sheetnames:
        ws = wbk["汇率"]
        for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            if row is None or not any(v not in (None, "") for v in row):
                continue
            vals = list(row) + [None] * 3
            mk = str(vals[0] or "").strip()
            cur = (str(vals[1] or "").strip().upper() or "USD")
            rate = _num_or_none(vals[2])
            if not _re.match(_MONTH_RE_STR, mk):
                errors.append(f"汇率 第{i}行：月份「{mk}」格式应为 YYYY-MM")
                continue
            if rate is None or rate <= 0:
                errors.append(f"汇率 第{i}行：汇率必须是正数")
                continue
            exrates.setdefault(mk, {})[cur] = rate

    wbk.close()
    return prices, exrates, pending, errors


@admin_bp.route("/api/admin/price/import", methods=["POST"])
def api_price_import_preview():
    """上传模板 → 只做解析和差异预览，不落盘正式配置。"""
    if not _check_user(allow_limited=True):
        return _forbid()
    f = request.files.get("file")
    if not f:
        return jsonify({"ok": False, "msg": "没有收到文件"})
    try:
        prices, exrates, pending, errors = _parse_price_workbook(f.read())
    except Exception as e:
        return jsonify({"ok": False, "msg": "解析 Excel 失败：" + str(e)})
    cfg = _load_price_cfg()
    diff = []
    for mk in sorted(prices.keys()):
        old = cfg.get("prices", {}).get(mk, {}) or {}
        new = prices[mk]
        added = sorted(set(new) - set(old))
        removed = sorted(set(old) - set(new))
        changed = sorted(k for k in set(new) & set(old) if new[k] != old[k])
        diff.append({"month": mk, "total": len(new),
                     "added": len(added), "changed": len(changed), "removed": len(removed),
                     "added_keys": added[:20], "changed_keys": changed[:20], "removed_keys": removed[:20]})
    ex_diff = []
    for mk in sorted(exrates.keys()):
        for cur, rate in exrates[mk].items():
            old_rate = cfg.get("exrates", {}).get(mk, {}).get(cur)
            if old_rate != rate:
                ex_diff.append({"month": mk, "currency": cur, "old": old_rate, "new": rate})
    # 暂存待确认导入的数据
    atomic_write_json(PRICE_PENDING, {"prices": prices, "exrates": exrates,
                                       "at": datetime.now().strftime("%Y-%m-%d %H:%M:%S")})
    return jsonify({"ok": True, "diff": diff, "ex_diff": ex_diff,
                    "pending": pending, "errors": errors,
                    "note": "导入按月整体覆盖：文件里出现的月份会替换该月全部价格；未出现的月份不受影响。"})


@admin_bp.route("/api/admin/price/import/confirm", methods=["POST"])
def api_price_import_confirm():
    if not _check_user(allow_limited=True):
        return _forbid()
    if not os.path.exists(PRICE_PENDING):
        return jsonify({"ok": False, "msg": "没有待确认的导入，请先上传预览"})
    with open(PRICE_PENDING, "r", encoding="utf-8") as fp:
        pend = json.load(fp)
    cfg = _load_price_cfg()
    months = []
    for mk, entries in (pend.get("prices") or {}).items():
        cfg.setdefault("prices", {})[mk] = entries   # 按月整体覆盖
        months.append(mk)
    for mk, rates in (pend.get("exrates") or {}).items():
        cfg.setdefault("exrates", {}).setdefault(mk, {}).update(rates)
    _save_price_cfg(cfg)
    try:
        os.remove(PRICE_PENDING)
    except Exception:
        pass
    return jsonify({"ok": True, "months": sorted(months)})


@admin_bp.route("/api/admin/price/recalc", methods=["POST"])
def api_price_recalc():
    """重算某月：只覆盖「算价/批量算价/重算」写入的行；单价为空的行顺带补算；手填价格一律不动。"""
    if not _check_user(allow_limited=True):
        return _forbid()
    import sqlite3
    import re as _re
    d = request.get_json(force=True, silent=True) or {}
    month = str(d.get("month", "")).strip()
    if not _re.match(_MONTH_RE_STR, month):
        return jsonify({"ok": False, "msg": "月份格式应为 YYYY-MM"})
    cfg = _load_price_cfg()
    conn = sqlite3.connect(YXO_DB)
    conn.row_factory = sqlite3.Row
    rows = conn.execute(
        'SELECT id, "口岸", "目的站", "箱属", "发班时间", "单箱价格", updated_by '
        'FROM records WHERE COALESCE(is_deleted,0)=0').fetchall()
    MACHINE = ("算价", "批量算价", "重算")
    scanned = updated = filled = failed = 0
    now = datetime.now().strftime("%Y-%m-%d %H:%M")
    for r in rows:
        mk = _mk_of_depart(r["发班时间"])
        if mk != month:
            continue
        scanned += 1
        old_raw = r["单箱价格"]
        is_empty = old_raw is None or str(old_raw).strip() == ""
        is_machine = (r["updated_by"] or "") in MACHINE
        if not (is_empty or is_machine):
            continue   # 手填价格不动
        p = _compute_price_local(r["口岸"], r["目的站"], r["箱属"], mk, cfg)
        if p is None:
            failed += 1
            continue
        old_num = _num_or_none(old_raw)
        if old_num is not None and abs(old_num - p) < 0.005:
            continue   # 没变化
        conn.execute('UPDATE records SET "单箱价格"=?, updated_at=?, updated_by=? WHERE id=?',
                     (p, now, "重算", r["id"]))
        if is_empty:
            filled += 1
        else:
            updated += 1
    if updated or filled:
        conn.execute(
            "INSERT INTO meta_kv (key, value) VALUES ('data_version', '1') "
            "ON CONFLICT(key) DO UPDATE SET value = CAST(CAST(value AS INTEGER) + 1 AS TEXT)")
    conn.commit()
    conn.close()
    return jsonify({"ok": True, "scanned": scanned, "updated": updated,
                    "filled": filled, "failed": failed})
