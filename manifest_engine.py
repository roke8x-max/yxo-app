# -*- coding: utf-8 -*-
"""舱单 / 箱号 统一导入引擎（芙蕾雅《舱单统一导入更新总体设计 v1.0》）。

流程：文件识别 → 归一化 → 以客编为键匹配库内 → 出差异清单(空跑) → 一键应用(快照+事务+留痕)。

铁律（违反即返工）：
  1. 七个字段永不碰（品名/放箱地/还箱/前端运输/后端运输/atb/运输条款）——它们根本不在 records 表列里，
     本引擎只写 MANIFEST_WRITABLE 白名单，故天然碰不到。
  2. 客编是核心键，箱号是复合兜底；箱号也定不了 → 报警，不猜。
  3. 对不上就报警，绝不自动写（除「空 → 有」补录）。
  4. 写库前必整库快照；应用失败整体回滚。

日期说明：设计文档原写「斜杠 YYYY/MM/DD」，但体检C已把库标准统一为 YYYY-MM-DD（横杠）并修复前端排序，
故本引擎发班时间也归一为 YYYY-MM-DD，与设计「语义无变化」判定一致，仅存储格式随现行库标准。
"""
import os
import re
import json
import sqlite3
from datetime import datetime, timedelta

from config import (
    PORT_ALIAS_LOOKUP, PORT_ALIASES, SUFFIX_DEST_MAP,
    MANIFEST_WRITABLE, ALL_FIELDS,
)
import openpyxl

BACKUP_DIR = r"D:\YXO_DATA\备份\数据库"


# ==================== 归一化 ====================

def code_core(code):
    """客编 core：去掉末尾 -字母 后缀，转大写。仅用于匹配，不写库。"""
    if not code:
        return ""
    c = str(code).strip().upper()
    return re.sub(r'-[A-Za-z]+$', '', c)


def code_suffix(code):
    """客编后缀（含前导 -）；无则空串。"""
    if not code:
        return ""
    m = re.search(r'(-[A-Za-z]+)$', str(code).strip())
    return m.group(1) if m else ""


def norm_box(b):
    return str(b).strip().upper() if b is not None else ""


def norm_train(raw):
    """YXO-2026-643 -> WB643；已是 WB 开头则 trim+大写；去前导零。"""
    if not raw:
        return ""
    s = str(raw).strip().upper()
    m = re.match(r'^YXO-?(\d{4})-?0*(\d+)$', s)
    if m:
        return f"WB{int(m.group(2))}"
    m2 = re.match(r'^(?:WB)?0*(\d+)$', s)
    return f"WB{int(m2.group(1))}" if m2 else s


def norm_port(raw):
    """舱单口岸别名 → 库标准写法；匹配不上返回 None（不写，报警）。"""
    if not raw:
        return None
    return PORT_ALIAS_LOOKUP.get(str(raw).strip().upper())


_DATE_PATS = ("%Y-%m-%d", "%Y/%m/%d", "%Y.%m.%d", "%Y年%m月%d日", "%m/%d/%Y", "%Y%m%d")


def norm_date_dash(raw):
    """任意日期 → YYYY-MM-DD（与现行库标准一致）。解析不了返回 None（不写，报警）。
    支持 Excel 日期序列号与 datetime.date/datetime 对象。"""
    if raw is None or raw == "":
        return None
    if isinstance(raw, bool):
        return None
    if isinstance(raw, (int, float)):
        try:
            return (datetime(1899, 12, 30) + timedelta(days=float(raw))).strftime("%Y-%m-%d")
        except Exception:
            return None
    if hasattr(raw, "strftime"):
        try:
            return raw.strftime("%Y-%m-%d")
        except Exception:
            return None
    s = str(raw).strip()
    # 预归一化分隔符：\ / . 统一转为 -（覆盖 Windows 反斜杠日期等边缘格式）
    s = re.sub(r'[\\/\.]', '-', s)
    for p in _DATE_PATS:
        try:
            return datetime.strptime(s, p).strftime("%Y-%m-%d")
        except ValueError:
            continue
    return None


def norm_seal(s):
    return str(s).strip() if s is not None else ""


def norm_owner(s):
    return str(s).strip().upper() if s is not None else ""


# ==================== 文件识别与解析 ====================

# 清洗正则：去掉空格/斜杠/反斜杠/括号(全半角)/连字符/下划线/冒号/中点/句号/全角空格等
# 非字母数字字符，转大写。用于模糊匹配表头，使「客户/编码」「箱 号」「客户编码（必填）」
# 都归一成「客户编码」，消除业务文件列名书写差异。
_CLEAN_RE = re.compile(r'[\s/\\()（）\-_：:·．.\u3000\uFF0D]+')


def clean_cell(s):
    if s is None:
        return ""
    return _CLEAN_RE.sub('', str(s)).upper()


COLUMN_MAP = {
    "客户编码": ["客户编码", "客编", "客户代码", "业务编号", "工作号", "工作单号",
                "订舱号", "订舱编号", "委托号", "S/O号", "SO号", "提单号", "运单号", "客户/编码"],
    "箱号": ["箱号", "箱号/封号", "container no", "container", "柜号", "箱柜号", "集装箱号"],
    "封号": ["封号", "客户自备封", "自备封", "seal", "铅封号", "封条号"],
    "箱属": ["箱属", "箱柜类型", "箱型", "箱柜", "箱种", "container type", "箱型尺寸"],
    "口岸": ["口岸", "口岸/车站", "起运口岸", "起运港", "port", "始发站", "发运站", "起运地"],
    "发班时间": ["班列确定发车时间", "发班时间", "发车时间", "班列确定发车日",
                "预计发车时间", "日期", "发班日期", "确定发车时间", "发车日", "发车日期", "班列日期"],
    "班列号": ["班列号", "班列", "train no", "train", "班列编号", "车次", "车号", "列车号"],
}


def _has_any(headers, names):
    """检查表头（清洗后）是否包含 names 中任意一个（子串匹配）。清洗消除空格/斜杠/括号差异。"""
    h = [clean_cell(x) for x in headers]
    return any(any(clean_cell(n) in hc for n in names) for hc in h)


def identify_file(headers):
    """用 COLUMN_MAP 别名全集做模糊识别，不硬编码具体名字。"""
    has_train = _has_any(headers, COLUMN_MAP["班列号"])
    has_code  = _has_any(headers, COLUMN_MAP["客户编码"])
    has_box   = _has_any(headers, COLUMN_MAP["箱号"])
    if has_train and has_code and has_box:
        return "manifest"
    if has_code and has_box and not has_train:
        return "box_template"
    # 宽松兜底：只要有客编就尝试按舱单解析（有些文件箱号列为空但结构是舱单）
    if has_code and has_train:
        return "manifest"
    # 再兜底：只有客编+日期/班列的清洗结果文件
    if has_code and (_has_any(headers, COLUMN_MAP["发班时间"]) or has_train):
        return "manifest"
    return "unknown"


def _score_header_row(row):
    """给一行打分：命中 COLUMN_MAP 关键字的列越多分越高，返回 (score, matched_set)。清洗后匹配。"""
    cells = [clean_cell(x) for x in row if x is not None and str(x).strip() != ""]
    score = 0
    matched = set()
    for tgt, names in COLUMN_MAP.items():
        for cell in cells:
            if any(clean_cell(n) in cell for n in names):
                if tgt not in matched:
                    matched.add(tgt)
                    score += 1
                break
    return score, matched


def _is_valid_header(matched):
    """合法表头：必须有客编，且（箱号 / 班列号 / 发班时间）至少一个。"""
    if "客户编码" not in matched:
        return False
    return bool(matched & {"箱号", "班列号", "发班时间"})


def _find_header_row(all_rows, look=60):
    """在前 look 行里找打分最高且合法的表头行；找不到则退回第 0 行。
    扫描范围放大到 60 行，覆盖封面/说明等多层前置标题（业务舱单常见）。"""
    best_idx, best_score = 0, 0
    for i, row in enumerate(all_rows[:look]):
        if row is None:
            continue
        if all(c is None or str(c).strip() == "" for c in row):
            continue  # 空行跳过
        score, matched = _score_header_row(row)
        if _is_valid_header(matched) and score > best_score:
            best_score, best_idx = score, i
    return best_idx


def _map_headers(headers):
    """列映射：精确优先，子串兜底；同一列不被两个目标重复占用。清洗后匹配。"""
    idx = {}
    used = set()
    h = [clean_cell(x) if x is not None else "" for x in headers]
    names_clean = {tgt: [clean_cell(n) for n in ns] for tgt, ns in COLUMN_MAP.items()}
    for tgt, ncl in names_clean.items():
        for i, cell in enumerate(h):
            if i in used:
                continue
            if cell in ncl:
                idx[tgt] = i
                used.add(i)
                break
    for tgt, ncl in names_clean.items():
        if tgt in idx:
            continue
        for i, cell in enumerate(h):
            if i in used:
                continue
            if any(nc and nc in cell for nc in ncl):
                idx[tgt] = i
                used.add(i)
                break
    return idx


def _pick_sheet(stream):
    """扫描所有 worksheet，返回 (ws_title, all_rows, sheet_names)。
    优先选第一个能识别为舱单/箱号模板的 sheet；若都识别不了，返回第一个 sheet（哪怕 unknown）。
    业务舱单常有封面 sheet，真实数据在后续 sheet 中。
    注意：必须用 read_only=False，因部分 xlsx 文件(如缺默认样式)在 read_only=True 下只读首列首行。"""
    wb = openpyxl.load_workbook(stream, read_only=False, data_only=True)
    try:
        sheets = wb.worksheets
        names = [s.title for s in sheets]
        for ws in sheets:
            all_rows = list(ws.iter_rows(values_only=True))
            if not all_rows:
                continue
            hdr_idx = _find_header_row(all_rows)
            if hdr_idx >= len(all_rows):
                continue
            if identify_file(list(all_rows[hdr_idx])) != "unknown":
                return ws.title, all_rows, names
        if sheets:
            return sheets[0].title, list(sheets[0].iter_rows(values_only=True)), names
        return None, [], names
    finally:
        wb.close()


def parse_workbook(stream):
    """从文件字节流解析 xlsx → (headers, rows, file_type)。支持多 sheet（优先识别第一个能认出类型的）。"""
    _, all_rows, _ = _pick_sheet(stream)
    if not all_rows:
        return [], [], "unknown"
    hdr_idx = _find_header_row(all_rows)
    headers = list(all_rows[hdr_idx])
    ftype = identify_file(headers)
    col = _map_headers(headers)
    rows = []
    for raw in all_rows[hdr_idx + 1:]:
        if raw is None:
            continue
        if all(c is None or str(c).strip() == "" for c in raw):
            continue
        rec = {}
        for tgt, ci in col.items():
            rec[tgt] = raw[ci] if ci < len(raw) else None
        if not rec.get("客户编码"):
            continue
        rows.append(rec)
    return headers, rows, ftype


def _diagnose_one(ws_title, all_rows):
    """对单个 sheet 的 all_rows 做完整诊断，返回 (headers, rows, ftype, diag_dict)。"""
    diag = {"steps": [], "error": None}
    total = len(all_rows)
    non_empty = sum(1 for r in all_rows if r is not None and any(c is not None and str(c).strip() != "" for c in r))
    diag["steps"].append({"name": "文件读取", "status": "OK",
                           "detail": f"[sheet:{ws_title}] 总{total}行, 非空{non_empty}行"})

    if not all_rows:
        diag["error"] = "文件为空（0行）"
        return [], [], "unknown", diag

    # 采样前10行原始内容（用于诊断）
    sample = []
    for i, row in enumerate(all_rows[:10]):
        cells = [(_truncate(str(c), 25) if c is not None else "") for c in (row or [])]
        sample.append(cells)
    diag["raw_sample"] = sample

    # 表头行定位
    hdr_idx = _find_header_row(all_rows)

    # 逐行打分详情（扫描范围与 _find_header_row 一致）
    scoring = []
    for i in range(min(60, len(all_rows))):
        row = all_rows[i]
        if row is None:
            continue
        if all(c is None or str(c).strip() == "" for c in row):
            scoring.append({"row": i, "score": None, "valid": None, "note": "空行"})
            continue
        score, matched = _score_header_row(row)
        valid = _is_valid_header(matched)
        selected = (i == hdr_idx)
        scoring.append({
            "row": i, "score": score, "valid": valid,
            "matched": sorted(matched),
            "preview": [_truncate(str(c), 20) for c in list(row)[:8]],
            "selected": selected,
        })
    diag["header_scoring"] = scoring
    diag["header_row_index"] = hdr_idx

    if hdr_idx >= len(all_rows):
        diag["error"] = f"表头定位越界: hdr_idx={hdr_idx} >= 总行数={total}"
        diag["steps"].append({"name": "表头定位", "status": "ERROR", "detail": diag["error"]})
        return [], [], "unknown", diag

    headers = list(all_rows[hdr_idx])
    h_display = [str(h)[:30] for h in headers]
    diag["steps"].append({
        "name": "表头定位",
        "status": "OK",
        "detail": f"第{hdr_idx}行, 列数={len(headers)}, 内容={h_display[:10]}"
    })
    diag["detected_headers"] = h_display

    # 文件类型识别
    ftype = identify_file(headers)
    col_check = {}
    for tgt, names in COLUMN_MAP.items():
        found = _has_any(headers, names)
        col_check[tgt] = {"found": found, "aliases_tried": names}
    diag["column_check"] = col_check
    diag["steps"].append({"name": "文件识别", "status": "OK" if ftype != "unknown" else "WARN",
                          "detail": f"type={ftype}"})

    # 列映射
    col = _map_headers(headers)
    diag["column_mapping"] = col  # {目标列名: 源索引}
    diag["steps"].append({"name": "列映射", "status": "OK" if len(col) >= 3 else "PARTIAL",
                          "detail": f"映射了{len(col)}列: {list(col.keys())}"})

    # 数据行解析
    rows = []
    skipped_empty = 0
    skipped_no_code = 0
    data_sample = []
    for ri, raw in enumerate(all_rows[hdr_idx + 1:], start=hdr_idx+1):
        if raw is None:
            continue
        if all(c is None or str(c).strip() == "" for c in raw):
            skipped_empty += 1
            continue
        rec = {}
        for tgt, ci in col.items():
            rec[tgt] = raw[ci] if ci < len(raw) else None
        if not rec.get("客户编码"):
            skipped_no_code += 1
            if len(data_sample) < 3:
                data_sample.append({
                    "row_idx": ri,
                    "reason": "客编为空",
                    "preview": [_truncate(str(c), 15) for c in (raw or [])[:5]],
                })
            continue
        rows.append(rec)
        if len(data_sample) < 3:
            data_sample.append({
                "row_idx": ri,
                "客户编码": str(rec.get("客户编码", ""))[:25],
                "箱号": str(rec.get("箱号", ""))[:15],
                "班列号": str(rec.get("班列号", ""))[:10],
            })

    diag["data_stats"] = {
        "parsed_rows": len(rows),
        "skipped_empty": skipped_empty,
        "skipped_no_code": skipped_no_code,
        "total_data_lines": total - hdr_idx - 1,
    }
    diag["data_sample"] = data_sample
    diag["steps"].append({
        "name": "数据行解析",
        "status": "OK" if len(rows) > 0 else "EMPTY",
        "detail": f"解析{len(rows)}行, 跳过空行{skipped_empty}, 客编为空{skipped_no_code}"
    })

    # 综合判断
    if ftype == "unknown":
        diag["error"] = "文件类型无法识别: 表头未匹配到舱单/箱号模板的关键字段组合（见下方列检查/表头打分）"
    elif len(rows) == 0:
        diag["error"] = f"无有效数据行: 共{total-hdr_idx-1}行数据, 但全部被跳过(空行{skipped_empty}+客编缺失{skipped_no_code})"

    return headers, rows, ftype, diag


def parse_workbook_diagnostics(stream):
    """带完整诊断的解析版本：返回 (headers, rows, ftype, diagnostics_dict)。
    支持多 sheet（优先识别第一个能认出舱单/箱号模板的 sheet）。"""
    diag = {"steps": [], "error": None}
    try:
        chosen_title, chosen_rows, sheet_names = _pick_sheet(stream)
    except Exception as e:
        diag["error"] = f"文件读取失败: {e}"
        diag["steps"].append({"name": "文件读取", "status": "ERROR", "detail": str(e)})
        return [], [], "unknown", diag

    multi_step = {"name": "多Sheet扫描", "status": "OK",
                  "detail": f"共{len(sheet_names)}个sheet({sheet_names})，使用「{chosen_title}」"}
    if chosen_title is None or not chosen_rows:
        diag["steps"].append(multi_step)
        diag["error"] = "文件无任何工作表或无法读取"
        return [], [], "unknown", diag

    # 对选定 sheet 做完整诊断
    headers, rows, ftype, sub = _diagnose_one(chosen_title, chosen_rows)
    for k, v in sub.items():
        diag[k] = v
    diag["steps"] = [multi_step] + (sub.get("steps") or [])
    return headers, rows, ftype, diag


def _truncate(s, maxlen):
    """截断字符串用于诊断显示。"""
    s = str(s) if s is not None else ""
    return (s[:maxlen-1] + "…") if len(s) > maxlen else s


def normalize_row(raw, ftype):
    """单行 → 归一化字典（仅含白名单字段 + 客编/core/suffix）。"""
    code = str(raw.get("客户编码") or "").strip()
    out = {
        "客户编码": code,
        "core": code_core(code),
        "suffix": code_suffix(code),
        "箱号": norm_box(raw.get("箱号")),
        "封号": norm_seal(raw.get("封号")),
        "箱属": norm_owner(raw.get("箱属")),
        "口岸": norm_port(raw.get("口岸")),
        "发班时间": norm_date_dash(raw.get("发班时间")),
        "班列号": norm_train(raw.get("班列号")),
    }
    return out


# ==================== 库内记录加载 ====================

def load_records(conn):
    rows = conn.execute(
        'SELECT id,"客户编码","箱号","班列号","口岸","发班时间","封号","箱属","目的站",'
        'COALESCE(is_deleted,0) AS del FROM records'
    ).fetchall()
    out = []
    for r in rows:
        code = r["客户编码"] or ""
        out.append({
            "id": r["id"], "code": code, "core": code_core(code),
            "suffix": code_suffix(code),
            "box": norm_box(r["箱号"]), "train": (r["班列号"] or "").strip(),
            "port": (r["口岸"] or "").strip(), "dep": (r["发班时间"] or "").strip(),
            "seal": (r["封号"] or "").strip(), "owner": (r["箱属"] or "").strip(),
            "dest": (r["目的站"] or "").strip(), "deleted": r["del"],
        })
    return out


# ==================== 差异构建（空跑，不写库）====================

def build_diff(conn, parsed_rows):
    """parsed_rows: list[normalized dict]。返回 {updates, imports, alerts, warnings}。"""
    recs = load_records(conn)
    active = [r for r in recs if not r["deleted"]]
    by_core = {}
    for r in active:
        by_core.setdefault(r["core"], []).append(r)
    by_train = {}
    for r in active:
        if r["train"]:
            by_train.setdefault(r["train"], []).append(r)

    updates, imports, alerts, warnings = [], [], [], []
    import_groups = {}   # train_no -> group dict
    seen_import_rows = set()

    for row in parsed_rows:
        core = row["core"]
        if not core:
            alerts.append(_alert("客编缺失", row, None, "客编解析为空，跳过"))
            continue
        cands = by_core.get(core, [])

        if not cands:
            # 整趟新专列？ → 看该班列号是否在库有其他箱
            if row["班列号"] and by_train.get(row["班列号"]):
                alerts.append(_alert("陌生客编", row, None,
                                     f"客编 {row['客户编码']} 库内无，但其班列 {row['班列号']} 已录过箱，疑似客编写错"))
                continue
            # 新专列待导入
            _add_import(import_groups, seen_import_rows, row)
            continue

        if len(cands) > 1:
            # 用箱号消歧
            box = row["箱号"]
            match = None
            if box:
                same = [c for c in cands if c["box"] == box]
                if len(same) == 1:
                    match = same[0]
            if match is None:
                alerts.append(_alert("客编重复无法定位", row, cands,
                                     f"客编 {row['客户编码']} 库内 {len(cands)} 行且箱号无法唯一确定"))
                continue
            db = match
        else:
            db = cands[0]
            # 唯一候选：箱号消歧校验
            if row["箱号"]:
                if db["box"] and db["box"] != row["箱号"]:
                    alerts.append(_alert("箱号冲突", row, db,
                                         f"客编 {row['客户编码']} 命中，但箱号 库内={db['box']} ≠ 文件={row['箱号']}"))
                    continue

        # 命中 → 计算字段差异
        changes = []
        # 非箱号字段（可覆盖：班列号/口岸/发班时间）
        for f in ("班列号", "口岸", "发班时间"):
            iv = row.get(f)
            if iv is None or iv == "":
                continue
            dv = db.get(f.lower()) or "" if f != "发班时间" else db["dep"]
            dv = db["train"] if f == "班列号" else (db["port"] if f == "口岸" else db["dep"])
            if dv == "":
                changes.append({"field": f, "old": "", "new": iv, "action": "补"})
            elif iv == dv:
                pass
            else:
                changes.append({"field": f, "old": dv, "new": iv, "action": "改"})
        # 箱号/封号/箱属（仅 空→有 补录；有→有不同 冲突）
        if row["箱号"]:
            for f, dkey in (("箱号", "box"), ("封号", "seal"), ("箱属", "owner")):
                iv = row.get(f)
                if iv is None or iv == "":
                    continue
                dv = db[dkey]
                if dv == "":
                    changes.append({"field": f, "old": "", "new": iv, "action": "补"})
                elif iv == dv:
                    pass
                else:
                    alerts.append(_alert("字段冲突", row, db,
                                         f"{f} 库内={dv} ≠ 文件={iv}，未自动写"))
                    continue
        # 后缀变更（专项）
        if db["suffix"] and row["suffix"] and db["suffix"] != row["suffix"]:
            alerts.append(_alert("后缀变更", row, db,
                                 f"客编后缀 库内={db['suffix']} → 文件={row['suffix']}",
                                 dest_suggest=SUFFIX_DEST_MAP.get(row["suffix"], "")))

        if changes:
            updates.append({
                "record_id": db["id"],
                "客户编码": db["code"],
                "箱号": db["box"],
                "changes": changes,
            })

    # 收尾 import groups
    for train_no, g in import_groups.items():
        if g["rows"]:
            imports.append(g)

    return {"updates": updates, "imports": imports, "alerts": alerts, "warnings": warnings}


def _alert(atype, row, db, msg, dest_suggest=""):
    a = {
        "type": atype,
        "客户编码": row["客户编码"],
        "箱号": row["箱号"],
        "文件值": {k: row.get(k) for k in ("班列号", "口岸", "发班时间", "箱属", "封号")},
        "说明": msg,
    }
    if isinstance(db, dict) and db:
        a["record_id"] = db["id"]
        a["库内值"] = {k: db[k] for k in ("train", "port", "dep", "box", "seal", "owner", "dest", "suffix")}
    elif isinstance(db, list) and db:
        # 多候选行（如客编重复）：取第一条的 id，库内值标明多候选
        a["record_id"] = db[0].get("id") if isinstance(db[0], dict) else None
        a["库内值"] = f"(共 {len(db)} 行候选，箱号无法消歧)"
    if dest_suggest:
        a["目的站建议"] = dest_suggest
    return a


def _add_import(groups, seen, row):
    tn = row["班列号"]
    if not tn:
        return
    g = groups.get(tn)
    if g is None:
        g = {
            "train_no": tn,
            "口岸": row["口岸"] or "",
            "发班时间": row["发班时间"] or "",
            "后缀": row["suffix"],
            "目的站建议": SUFFIX_DEST_MAP.get(row["suffix"], ""),
            "rows": [],
        }
        groups[tn] = g
    key = (row["客户编码"], row["箱号"])
    if key in seen:
        return
    seen.add(key)
    g["rows"].append({
        "客户编码": row["客户编码"],
        "箱号": row["箱号"],
        "封号": row["封号"],
        "箱属": row["箱属"],
        "口岸": row["口岸"] or g["口岸"],
        "发班时间": row["发班时间"] or g["发班时间"],
    })


# ==================== 应用（快照 + 事务 + 留痕）====================

def _snapshot_path():
    os.makedirs(BACKUP_DIR, exist_ok=True)
    return os.path.join(BACKUP_DIR, datetime.now().strftime("apply_%Y%m%d_%H%M.db"))


def _next_batch_id(conn):
    day = datetime.now().strftime("%Y%m%d")
    n = conn.execute(
        "SELECT COUNT(*) FROM import_batch WHERE batch_id LIKE ?", (f"IMP-{day}-%",)
    ).fetchone()[0]
    return f"IMP-{day}-{n + 1:03d}"


def apply_diff(conn, diff, operator, source_files):
    """diff: {updates:[{record_id, changes:[{field,new}]}], imports:[{train_no, 目的站, rows:[...]}],
               alerts_applied:[{record_id, new_code?, 目的站?}]}
    返回 batch_id；失败抛异常（调用方负责回滚/不提交）。"""
    snap = _snapshot_path()
    # 1. 整库快照（失败 → 抛异常中止，不写任何东西）
    # 先 checkpoint，确保 WAL 已并入主库文件，快照完整一致。
    try:
        conn.execute("PRAGMA wal_checkpoint(FULL)")
    except Exception:
        pass
    import shutil
    shutil.copyfile(conn.execute("PRAGMA database_list").fetchone()[2], snap)

    batch_id = _next_batch_id(conn)
    now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    n_update = n_insert = n_alert = 0

    # 2. 更新通道
    for u in diff.get("updates", []):
        rid = u["record_id"]
        for ch in u["changes"]:
            f = ch["field"]
            new_val = ch["new"]
            old = conn.execute(f'SELECT "{f}" FROM records WHERE id=?', (rid,)).fetchone()
            old_val = old[f] if old else ""
            if str(old_val) == str(new_val):
                continue
            conn.execute(
                f'UPDATE records SET "{f}"=?, updated_at=?, updated_by=? WHERE id=?',
                (new_val, now, operator, rid))
            conn.execute(
                "INSERT INTO update_log(batch_id,batch_type,record_id,客户编码,箱号,field,"
                "old_value,new_value,action,source_file,operator,created_at) "
                "VALUES(?,?,?,?,?,?,?,?,?,?,?,?)",
                (batch_id, "update", rid, u.get("客户编码", ""), u.get("箱号", ""),
                 f, old_val, new_val, ch.get("action", "改"),
                 ",".join(source_files), operator, now))
            n_update += 1

    # 3. 新专列导入通道
    for g in diff.get("imports", []):
        tn = g["train_no"]
        dest = g.get("目的站") or g.get("目的站建议") or ""
        month = (g.get("发班时间") or "")[:7]
        if month and re.match(r'^\d{4}-\d{2}$', month):
            pass
        else:
            month = ""
        max_seq = conn.execute("SELECT COALESCE(MAX(seq),0)+1 FROM records").fetchone()[0]
        for i, r in enumerate(g["rows"]):
            seq = max_seq + i
            cols = ["seq", "order_idx", "客户编码", "箱号", "封号", "箱属", "班列号",
                    "口岸", "发班时间", "目的站", "班列类型", "状态", "台账月份",
                    "is_deleted", "updated_by", "updated_at"]
            vals = [seq, float(seq), r.get("客户编码", ""), r.get("箱号", ""), r.get("封号", ""),
                    r.get("箱属", ""), tn, r.get("口岸", "") or g.get("口岸", ""),
                    r.get("发班时间", "") or g.get("发班时间", ""), dest, "专列", "正常",
                    month, 0, operator, now]
            ph = ",".join("?" * len(cols))
            csql = ",".join(f'"{c}"' for c in cols)
            cur = conn.execute(f"INSERT INTO records ({csql}) VALUES ({ph})", vals)
            new_id = cur.lastrowid
            conn.execute(
                "INSERT INTO update_log(batch_id,batch_type,record_id,客户编码,箱号,field,"
                "old_value,new_value,action,source_file,operator,created_at) "
                "VALUES(?,?,?,?,?,?,?,?,?,?,?,?)",
                (batch_id, "import", new_id, r.get("客户编码", ""), r.get("箱号", ""),
                 "(新增专列箱)", "", f"{r.get('客户编码','')}/{r.get('箱号','')}",
                 "增", ",".join(source_files), operator, now))
            n_insert += 1

    # 4. 后缀变更确认（alerts_applied）
    for a in diff.get("alerts_applied", []):
        rid = a["record_id"]
        sets, params = [], []
        if a.get("new_code"):
            sets.append('"客户编码"=?'); params.append(a["new_code"])
        if a.get("目的站"):
            sets.append('"目的站"=?'); params.append(a["目的站"])
        if sets:
            old = conn.execute(f'SELECT "客户编码","目的站" FROM records WHERE id=?', (rid,)).fetchone()
            conn.execute(
                f'UPDATE records SET {",".join(sets)}, updated_at=?, updated_by=? WHERE id=?',
                params + [now, operator, rid])
            conn.execute(
                "INSERT INTO update_log(batch_id,batch_type,record_id,客户编码,箱号,field,"
                "old_value,new_value,action,source_file,operator,created_at) "
                "VALUES(?,?,?,?,?,?,?,?,?,?,?,?)",
                (batch_id, "update", rid, old["客户编码"] if old else "", "",
                 "客户编码/目的站", "", a.get("new_code", "") + ("|" + a.get("目的站", ""))[:0],
                 "改", ",".join(source_files), operator, now))
            n_alert += 1

    # 5. 批次记录
    conn.execute(
        "INSERT INTO import_batch(batch_id,batch_type,source_files,snapshot,"
        "n_update,n_insert,n_alert,operator,created_at) VALUES(?,?,?,?,?,?,?,?,?)",
        (batch_id, "manifest", ",".join(source_files), snap, n_update, n_insert, n_alert,
         operator, now))

    return batch_id


# ==================== 回退 / 还原 ====================

def revert_batch(conn, batch_id):
    """整批回退：update 类反向写回 old；import 类软删除。标记 reverted=1。"""
    logs = conn.execute(
        "SELECT id,record_id,field,old_value,action,batch_type FROM update_log "
        "WHERE batch_id=? AND COALESCE(reverted,0)=0", (batch_id,)
    ).fetchall()
    now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    for lg in logs:
        if lg["batch_type"] == "import":
            conn.execute("UPDATE records SET is_deleted=1, deleted_at=?, deleted_by=? WHERE id=?",
                         (now, "manifest回退", lg["record_id"]))
        else:
            if lg["field"] in ("客户编码", "目的站") or lg["field"] == "(新增专列箱)":
                continue
            conn.execute(f'UPDATE records SET "{lg["field"]}"=? WHERE id=?',
                         (lg["old_value"], lg["record_id"]))
        conn.execute("UPDATE update_log SET reverted=1, reverted_at=? WHERE id=?",
                     (now, lg["id"]))
    conn.execute("UPDATE import_batch SET reverted=1 WHERE batch_id=?", (batch_id,))


def revert_item(conn, log_id):
    """单条撤销。"""
    lg = conn.execute(
        "SELECT id,record_id,field,old_value,action,batch_type,batch_id FROM update_log WHERE id=?",
        (log_id,)).fetchone()
    if not lg or lg["reverted"]:
        return False
    now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    if lg["batch_type"] == "import":
        conn.execute("UPDATE records SET is_deleted=1, deleted_at=?, deleted_by=? WHERE id=?",
                     (now, "manifest回退", lg["record_id"]))
    else:
        if lg["field"] in ("客户编码", "目的站"):
            pass
        else:
            conn.execute(f'UPDATE records SET "{lg["field"]}"=? WHERE id=?',
                         (lg["old_value"], lg["record_id"]))
    conn.execute("UPDATE update_log SET reverted=1, reverted_at=? WHERE id=?", (now, lg["id"]))
    # 若批次内全部已回退，标记批次
    left = conn.execute("SELECT COUNT(*) FROM update_log WHERE batch_id=? AND COALESCE(reverted,0)=0",
                        (lg["batch_id"],)).fetchone()[0]
    if left == 0:
        conn.execute("UPDATE import_batch SET reverted=1 WHERE batch_id=?", (lg["batch_id"],))
    return True


def restore_snapshot(conn, batch_id):
    """核弹级：用批次快照整库还原。调用方需二次确认。"""
    row = conn.execute("SELECT snapshot FROM import_batch WHERE batch_id=?", (batch_id,)).fetchone()
    if not row or not row["snapshot"]:
        raise ValueError("无快照文件")
    snap = row["snapshot"]
    if not os.path.exists(snap):
        raise ValueError("快照文件不存在: " + snap)
    db_path = conn.execute("PRAGMA database_list").fetchone()[2]
    # 先关闭连接释放文件句柄，避免 Windows 下复制时句柄冲突 / 视图失效。
    conn.close()
    import shutil
    shutil.copyfile(snap, db_path)
    return snap
