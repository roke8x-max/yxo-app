# -*- coding: utf-8 -*-
"""
从 Excel 导入 / 更新数据到 SQLite。
- 列头用"关键词匹配"，兼容 8 月标准列与 7 月历史台账（散舱）等不同列名。
- 新行（按"客户编码"判断）：插入，操作字段套用默认值。
- 已存在的行：只刷新"基础字段"（来自 Excel），保留同事已填的操作字段。
- 特殊：台账里"入堆场/入站时间"为合并列，其右侧紧邻的无表头列是入站时间。
- 可被 app.py 调用：from import_excel import run_import
- 也可命令行单独跑：python import_excel.py [filepath]
"""
import os
import re
import sys
import sqlite3
import datetime as _dt
import config

# 日期类型字段（导入时统一格式为 YYYY/MM/DD）
_DATE_FIELDS = {f["name"] for f in config.FIELD_DEFS if f.get("type") == "date"}
# 新增 / 导入时自动填充的默认值
DEFAULTS = {f["name"]: f["default"] for f in config.FIELD_DEFS if f.get("default")}

# 表头关键词 -> 字段名（按顺序匹配，先命中先占）
HEADER_RULES = [
    ("客户编码", "客户编码"),
    ("箱号", "箱号"),
    ("封号", "封号"),
    ("箱属", "箱属"),
    ("发班", "发班时间"),
    ("班列", "班列号"),
    ("口岸", "口岸"),
    ("目的站", "目的站"),
    ("到站", "目的站"),
    ("入堆场", "入堆场"),       # 命中后，右侧紧邻无表头列 -> 入站
    ("随车", "随车"),
    ("草单", "草单"),
    ("报放", "报放单"),
    ("开票", "开票子公司名称"),
    ("子公司", "开票子公司名称"),
    ("货源类型", "货源类型"),
    ("本地货源", "本地货源公司"),
    ("备注", "备注"),
    ("价格", "单箱价格"),
    ("人民币", "单箱价格"),
]

# 值归一化：台账里的简写 -> 标准选项
VALUE_MAP = {
    "随车": {"已发": "已发邮件"},
}


def _build_col_map(header):
    """根据表头建立 列索引 -> 字段名 的映射"""
    col_map = {}
    used = set()
    for i, h in enumerate(header):
        if h is None:
            continue
        h = str(h).strip()
        for key, field in HEADER_RULES:
            if key in h and field not in used:
                col_map[i] = field
                used.add(field)
                # 入堆场/入站时间 合并列：右侧紧邻无表头列视为入站
                if key == "入堆场" and (i + 1) < len(header) and header[i + 1] in (None, ""):
                    if "入站" not in used:
                        col_map[i + 1] = "入站"
                        used.add("入站")
                break
    return col_map


def _norm_date(raw, year, dep_date=None):
    """日期字段规范化：datetime/date -> YYYY/MM/DD；文本 '6-29入站' -> YYYY/MM/DD

    跨年修正：无年份文本用发班年补全后，若与发班日期相差 300 天以上，
    说明跨了年（如 12 月发班、次年 1 月入站），自动把年份 ±1 修正。"""
    if raw is None:
        return ""
    if isinstance(raw, (_dt.datetime, _dt.date)):
        return raw.strftime("%Y/%m/%d")
    s = str(raw).strip()
    if not s:
        return ""
    # 已经是完整日期串
    for fmt in ("%Y/%m/%d", "%Y-%m-%d", "%Y/%m/%d %H:%M:%S", "%Y-%m-%d %H:%M:%S"):
        try:
            return _dt.datetime.strptime(s[:19], fmt).strftime("%Y/%m/%d")
        except Exception:
            continue
    # 文本形如 '6-29入站' / '7-1入站' -> 提取月-日，用发班年份补全年
    m = re.search(r"(\d{1,2})\s*[-/]\s*(\d{1,2})", s)
    if m:
        try:
            mo, dd = int(m.group(1)), int(m.group(2))
            y = year
            if dep_date is not None:
                cand = _dt.date(y, mo, dd)
                if (dep_date - cand).days > 300:      # 补出的日期比发班早太多 → 实际是次年
                    y += 1
                elif (cand - dep_date).days > 300:    # 补出的日期比发班晚太多 → 实际是上一年
                    y -= 1
            return f"{y}/{mo:02d}/{dd:02d}"
        except Exception:
            pass
    return s  # 实在解析不出就原样保留（不丢数据）


def _norm_select(field, raw):
    if raw is None:
        return ""
    s = str(raw).strip()
    return VALUE_MAP.get(field, {}).get(s, s)


def analyze_import(filepath):
    """预检：统计文件里各班列号的票数，用于散舱/专列自动预判。
    返回 {班列号: 票数}；读不出来返回 {}。"""
    try:
        import openpyxl
        wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        wb.close()
    except Exception:
        return {}
    if not rows:
        return {}
    col_map = _build_col_map(rows[0])
    train_idx = next((i for i, f in col_map.items() if f == "班列号"), None)
    if train_idx is None:
        return {}
    counts = {}
    for r in rows[1:]:
        if r is None or train_idx >= len(r):
            continue
        t = str(r[train_idx] or "").strip()
        if t:
            counts[t] = counts.get(t, 0) + 1
    return counts


def run_import(conn, filepath=None, train_type=None):
    """train_type: '散舱' / '专列'，导入时写入班列类型；None 则新行默认散舱、旧行不动。"""
    fp = filepath or config.IMPORT_FILE
    if not os.path.exists(fp):
        print(f"[导入跳过] 文件不存在: {fp}")
        return 0
    try:
        import openpyxl
        wb = openpyxl.load_workbook(fp, read_only=True, data_only=True)
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        wb.close()
    except Exception as e:
        print(f"[导入失败] 读取 Excel 出错: {e}")
        return 0
    if not rows:
        return 0

    header = [str(h).strip() if h is not None else "" for h in rows[0]]
    col_map = _build_col_map(rows[0])
    print(f"[导入] 列映射: {col_map}")

    n_new, n_upd = 0, 0
    for r in rows[1:]:
        if r is None:
            continue
        # 组装 {字段: 原始值}
        raw = {}
        for i, field in col_map.items():
            if i < len(r):
                raw[field] = r[i]
        if not any(v not in (None, "") for v in raw.values()):
            continue
        # 发班年份 + 发班日期（用于文本日期补全年和跨年修正）
        year = _dt.date.today().year
        dep_date = None
        dep = raw.get("发班时间")
        if isinstance(dep, _dt.datetime):
            year, dep_date = dep.year, dep.date()
        elif isinstance(dep, _dt.date):
            year, dep_date = dep.year, dep
        elif isinstance(dep, str):
            m = re.search(r"(\d{4})", dep)
            if m:
                year = int(m.group(1))
            m2 = re.match(r"^\s*(\d{4})[-/](\d{1,2})[-/](\d{1,2})", dep)
            if m2:
                try:
                    dep_date = _dt.date(int(m2.group(1)), int(m2.group(2)), int(m2.group(3)))
                except Exception:
                    pass
        # 规范化
        rec = {}
        for field, v in raw.items():
            if field in _DATE_FIELDS:
                rec[field] = _norm_date(v, year, dep_date)
            elif field in VALUE_MAP:
                rec[field] = _norm_select(field, v)
            else:
                rec[field] = "" if v is None else str(v)

        # 班列号规范化：与既有数据约定一致，纯数字 / 小写 wb 都统一为「WB + 数字」
        # （如 1234 → WB1234、wb759 → WB759），使无前缀的 4 位数也能被机器人按班列号查到
        _t = (rec.get("班列号") or "").strip()
        if _t:
            _tm = re.match(r"^(?:wb\s*)?(\d{2,6})$", _t, re.IGNORECASE)
            if _tm:
                rec["班列号"] = "WB" + _tm.group(1)

        key = rec.get("客户编码", "")
        # 台账月份：默认随发班时间所在月（仅在为空时填，不覆盖用户后续手动调整）
        lm = rec.get("台账月份", "")
        if not lm or not str(lm).strip():
            dep = rec.get("发班时间", "")
            if dep:
                m = str(dep)[:7].replace("/", "-")
                if re.match(r"^\d{4}-\d{2}$", m):
                    rec["台账月份"] = m
        exist = conn.execute('SELECT id FROM records WHERE "客户编码"=?', (key,)).fetchone() if key else None
        if exist:
            sets = ", ".join([f'"{f}"=?' for f in config.BASE_FIELDS if f in rec])
            conn.execute(f"UPDATE records SET {sets} WHERE id=?",
                         [rec.get(f, "") for f in config.BASE_FIELDS if f in rec] + [exist["id"]])
            # 用户显式选了类型 → 本次导入涉及的旧行也同步该类型
            if train_type:
                conn.execute('UPDATE records SET "班列类型"=? WHERE id=?',
                             (train_type, exist["id"]))
            n_upd += 1
        else:
            seq = conn.execute("SELECT COALESCE(MAX(seq),0)+1 AS s FROM records").fetchone()["s"]
            cols = config.ALL_FIELDS
            ph = ", ".join(["?"] * len(cols))
            col_sql = ", ".join([f'"{c}"' for c in cols])
            vals = [rec.get(c, "") for c in cols]
            cur = conn.execute(f"INSERT INTO records (seq, {col_sql}) VALUES (?, {ph})", [seq] + vals)
            # 班列类型不在 ALL_FIELDS（页面暂不显示），单独写入
            conn.execute('UPDATE records SET "班列类型"=? WHERE id=?',
                         (train_type or "散舱", cur.lastrowid))
            n_new += 1
    conn.commit()
    print(f"[导入] {os.path.basename(fp)} 新增 {n_new} 行，更新 {n_upd} 行")
    return n_new + n_upd


if __name__ == "__main__":
    os.makedirs(config.DATA_DIR, exist_ok=True)
    fp = sys.argv[1] if len(sys.argv) > 1 else None
    c = sqlite3.connect(config.DB_PATH)
    c.row_factory = sqlite3.Row
    run_import(c, fp)
    c.close()
