# -*- coding: utf-8 -*-
"""
托书生成引擎（原型 / 可集成版）  2026-07-31 芙蕾雅设计
=====================================================
设计目标：把「托书自动生成」从"硬编码单元格 + 手写字典 + 手动输日期"的桌面脚本，
改造成「模板字段映射 + 规则配置 + 数据源查询」的可维护模块。

本文件是纯逻辑模块（无 Flask 依赖），可由 yxo_app 直接 import：
    from tuoshu_engine import TuoshuEngine, build_default_config

数据源：调用方传入「订舱记录列表」（结构同 yxo.db 的 records 表查询结果），
        而非再去读一个手维护的「舱单列表.xlsx」。
模板：模板文件存放在 yxo_app/data/tuoshu_templates/，字段映射写在模板配置里。
"""
import copy
import datetime
import os
import re

import openpyxl


# ---------------------------------------------------------------------------
# 目的站 → 英文/国家 解析表（替代原脚本里按"客户编码后缀"硬匹配的 dest_mapping）
# 键用系统「目的站」字段的干净中文值；值 = 托书 E15 需要的 "EN中文, COUNTRY" 串。
# 这套表应落到系统「选项维护 / 托书目的站映射」里，由用户自助增删，无需改代码。
# ---------------------------------------------------------------------------
DEST_MAP = {
    "沃尔西诺": "Vorsino沃尔西诺, RUS",
    "科里亚季奇": "Kolyadichi科里亚季奇, BEL",
    "叶卡捷琳堡": "Yekaterinburg叶卡捷琳堡, RUS",
    "克列西哈": "Kleshchikha克列西哈, RUS",
    "电煤": "Dostyk电煤, KAZ",
    "杜伊斯堡": "Duisburg杜伊斯堡, GER",
    "杜伊斯堡（时刻表）": "Duisburg杜伊斯堡, GER",
    "布达佩斯": "Budapest布达佩斯, HUN",
    "马拉": "Malaszewicze马拉, POL",
    "马拉（时刻表）": "Malaszewicze马拉, POL",
    "别雷拉斯特": "Bely Rast别雷拉斯特, RUS",
    "谢利亚季诺": "Selyatino谢利亚季诺, RUS",
    "谢丽亚基诺": "Selyatino谢利亚季诺, RUS",
    "莫斯科": "Moscow莫斯科, RUS",
    "明斯克": "Minsk明斯克, BLR",
    "满洲里": "Manzhouli满洲里, CN",
    "罗斯托夫": "Rostov罗斯托夫, RUS",
}

# 目的站同义归一（系统里存在错别字/带后缀的变体，先归一成上表的键）
DEST_NORMALIZE = {
    "谢丽亚基诺": "谢利亚季诺",
    "杜伊斯堡（时刻表）": "杜伊斯堡",
    "马拉（时刻表）": "马拉",
}


def resolve_destination(station_cn: str) -> str:
    """把系统「目的站」解析成托书 E15 需要的字符串。"""
    if not station_cn:
        return "未知目的站"
    s = DEST_NORMALIZE.get(station_cn, station_cn)
    return DEST_MAP.get(s, f"未知目的站({station_cn})")


def train_no_suffix(train_no: str) -> str:
    """班列号取末段，用于文件名，如 'YXE-20260718-653' -> '653'。"""
    if not train_no:
        return ""
    return str(train_no).split("-")[-1]


def extract_cn(text: str) -> str:
    m = re.search(r"[\u4e00-\u9fff]+", text or "")
    return m.group(0) if m else ""


# ---------------------------------------------------------------------------
# 默认模板字段映射（decouple：模板换了只改这里，不改逻辑）
# 逻辑字段名 -> (工作表名, 单元格)
# ---------------------------------------------------------------------------
DEFAULT_FIELD_MAP = {
    "booking_date":   ("Sheet1", "E4"),    # 订舱日期
    "departure_date": ("Sheet1", "E13"),   # 发运日期
    "destination":    ("Sheet1", "E15"),   # 目的站
    "container":      ("Sheet1", "B24"),   # 数量和箱型
}

# 默认命名规则（可用占位符，后续放系统配置）
DEFAULT_NAME_PATTERN = "渝新欧订舱委托书 -{departure:%Y%m%d} -{station_cn} -{train_suffix}"


class TuoshuEngine:
    def __init__(self, template_path: str, field_map: dict = None,
                 name_pattern: str = DEFAULT_NAME_PATTERN, dest_map: dict = None,
                 dest_normalize: dict = None):
        self.template_path = template_path
        self.field_map = field_map or DEFAULT_FIELD_MAP
        self.name_pattern = name_pattern
        self.dest_map = dest_map              # 注入覆盖（系统表 tuoshu_dest_map）
        self.dest_normalize = dest_normalize  # 同义归一（系统表 alias 列）

    # ---- 目的站解析：优先用注入的系统表，回退模块内置表 ----
    # 2026-08-03 小叽集成修复：原 self.dest_map 只存不用，注入永远不生效。
    def _resolve(self, station_cn: str) -> str:
        if not self.dest_map:
            return resolve_destination(station_cn)
        if not station_cn:
            return "未知目的站"
        s = (self.dest_normalize or {}).get(station_cn, station_cn)
        return self.dest_map.get(s, f"未知目的站({station_cn})")

    # ---- 单条班列的生成 ----
    def generate_one(self, group: dict, booking_date: str, output_dir: str) -> str:
        """
        group: {
            "train_no": "YXE-20260718-653",
            "departure_date": "2026-07-18",     # 字符串 YYYY-MM-DD
            "station_cn": "沃尔西诺",            # 系统「目的站」干净值
            "boxes": 12,                         # 箱量（已按班列聚合）
            "container_str": "12*40'HQ",        # 可选：自定义箱型串
        }
        booking_date: 订舱日期字符串 YYYY-MM-DD
        """
        wb = openpyxl.load_workbook(self.template_path)
        sheet_name, cell = self.field_map["booking_date"]
        wb[sheet_name][cell] = booking_date
        _, cell = self.field_map["departure_date"]
        wb[sheet_name][cell] = group["departure_date"]
        _, cell = self.field_map["destination"]
        wb[sheet_name][cell] = self._resolve(group.get("station_cn"))
        _, cell = self.field_map["container"]
        wb[sheet_name][cell] = group.get("container_str") or f"{group['boxes']}*40'HQ"

        departure = datetime.datetime.strptime(group["departure_date"], "%Y-%m-%d")
        station_cn = extract_cn(self._resolve(group.get("station_cn")))
        fname = self.name_pattern.format(
            departure=departure,
            station_cn=station_cn or group.get("station_cn", ""),
            train_suffix=train_no_suffix(group.get("train_no", "")),
        )
        os.makedirs(output_dir, exist_ok=True)
        out = os.path.join(output_dir, fname + ".xlsx")
        wb.save(out)
        return out

    # ---- 批量：按班列号聚合记录并生成 ----
    def generate_batch(self, records: list, booking_date: str, output_dir: str) -> list:
        """
        records: yxo.db records 查询结果列表，每条含
            班列号 / 发班时间 / 目的站 / 箱号（可选箱属）
        按「班列号」聚合，自动算箱量。
        """
        groups = {}
        for r in records:
            tno = r.get("班列号") or r.get("train_no") or "未知班列"
            g = groups.setdefault(tno, {
                "train_no": tno,
                "departure_date": (r.get("发班时间") or r.get("departure_date") or "")[:10],
                "station_cn": r.get("目的站") or r.get("station_cn") or "",
                "boxes": 0,
            })
            if (r.get("箱号") or r.get("box_no")):
                g["boxes"] += 1
        out_files = []
        for g in groups.values():
            if g["boxes"] == 0:
                g["boxes"] = 1  # 至少 1 箱，避免 0*40'HQ
            out_files.append(self.generate_one(g, booking_date, output_dir))
        return out_files


# ---------------------------------------------------------------------------
# 小工具：从 yxo.db 的 records 表直接取数（示范；实际由 app 路由调用）
# ---------------------------------------------------------------------------
def fetch_records_by_train(db_path: str, train_no: str = None,
                           date_from: str = None, date_to: str = None) -> list:
    import sqlite3
    conn = sqlite3.connect(db_path)
    conn.row_factory = sqlite3.Row
    sql = 'SELECT "班列号","发班时间","目的站","箱号" FROM records WHERE COALESCE(is_deleted,0)=0'
    params = []
    if train_no:
        sql += ' AND "班列号"=?'; params.append(train_no)
    if date_from:
        sql += ' AND "发班时间">=?'; params.append(date_from)
    if date_to:
        sql += ' AND "发班时间"<=?'; params.append(date_to)
    rows = [dict(x) for x in conn.execute(sql, params).fetchall()]
    conn.close()
    return rows


if __name__ == "__main__":
    # 用真实模板 + 模拟记录跑一遍，验证字段映射正确
    here = os.path.dirname(os.path.abspath(__file__))
    tpl = os.path.join(here, "托书模板.xlsx")
    out_dir = os.path.join(here, "输出验证")
    eng = TuoshuEngine(tpl)
    demo = [
        {"班列号": "YXE-20260718-653", "发班时间": "2026-07-18", "目的站": "沃尔西诺", "箱号": "ABC1"},
        {"班列号": "YXE-20260718-653", "发班时间": "2026-07-18", "目的站": "沃尔西诺", "箱号": "ABC2"},
        {"班列号": "YXE-20260720-710", "发班时间": "2026-07-20", "目的站": "电煤", "箱号": "DEF1"},
    ]
    files = eng.generate_batch(demo, "2026-07-31", out_dir)
    # 回读校验
    for f in files:
        wb = openpyxl.load_workbook(f)
        ws = wb["Sheet1"]
        print(f"✓ {os.path.basename(f)}")
        print(f"    E4(订舱日期)={ws['E4'].value}  E13(发运)={ws['E13'].value}  "
              f"E15(目的站)={ws['E15'].value}  B24(箱量)={ws['B24'].value}")
